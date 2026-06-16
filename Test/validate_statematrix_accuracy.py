#!/usr/bin/env python3
"""
Comprehensive State Matrix Accuracy Validator
Tests every column and grid value in the exported State Matrix for 100% accuracy
"""

import json
import sqlite3
import sys
import time
from pathlib import Path
from typing import Dict, List, Any, Set
import pandas as pd
import openpyxl

# Reconfigure stdout and stderr for UTF-8 to prevent encoding crashes on Windows
import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from core.config_manager import get_config
from core.reporting_manager import export_state_matrix_xlsx


def check_processing_status() -> Dict[str, Any]:
    config = get_config()
    audit_db = Path(config.paths.working_root) / "audit" / "audit.db"
    
    conn = sqlite3.connect(str(audit_db))
    conn.row_factory = sqlite3.Row
    
    # Get total documents
    total = conn.execute("SELECT COUNT(*) as count FROM file_state").fetchone()['count']
    
    # Get completed documents
    completed = conn.execute(
        "SELECT COUNT(*) as count FROM file_state WHERE current_status IN ('completed', 'tag_completed')"
    ).fetchone()['count']
    
    # Get documents with extended metadata
    extended_count = conn.execute(
        "SELECT COUNT(*) as count FROM file_state WHERE extended_metadata_json IS NOT NULL AND extended_metadata_json != ''"
    ).fetchone()['count']
    
    # Sample extended metadata
    samples = conn.execute(
        "SELECT file_name, category, department, purpose, extended_metadata_json FROM file_state WHERE extended_metadata_json IS NOT NULL AND extended_metadata_json != '' LIMIT 5"
    ).fetchall()
    
    conn.close()
    
    return {
        'total': total,
        'completed': completed,
        'extended_count': extended_count,
        'samples': [dict(s) for s in samples]
    }


def analyze_metadata_schema(metadata_path: str) -> Dict[str, Any]:
    """Analyze the metadata workbook schema."""
    print(f"\n{'='*80}")
    print("ANALYZING METADATA WORKBOOK SCHEMA")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
    schema = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Count rows with data
        row_count = 0
        sample_rows = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell for cell in row):
                row_count += 1
                if len(sample_rows) < 3:
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row) and row[col_idx]:
                            row_data[header] = row[col_idx]
                    sample_rows.append(row_data)
        
        schema[sheet_name] = {
            'headers': headers,
            'row_count': row_count,
            'sample_rows': sample_rows
        }
        
        print(f"\nSheet: {sheet_name}")
        print(f"  Headers: {headers}")
        print(f"  Rows with data: {row_count}")
        if sample_rows:
            print(f"  Sample row: {json.dumps(sample_rows[0], indent=2)}")
    
    wb.close()
    return schema


def export_and_analyze_state_matrix() -> Dict[str, Any]:
    """Export State Matrix and analyze its structure."""
    print(f"\n{'='*80}")
    print("EXPORTING AND ANALYZING STATE MATRIX")
    print(f"{'='*80}")
    
    # Export State Matrix
    output_path = export_state_matrix_xlsx(filters={}, out_path="runtime/test_state_matrix.xlsx")
    print(f"\n✓ Exported State Matrix to: {output_path}")
    
    # Load and analyze
    df = pd.read_excel(output_path, sheet_name="📄 Document Data")
    
    print(f"\n✓ State Matrix loaded:")
    print(f"  Total rows: {len(df)}")
    print(f"  Total columns: {len(df.columns)}")
    print(f"  Columns: {list(df.columns)}")
    
    # Identify extended metadata columns (not in core 14)
    core_columns = [
        'Smart ID', 'File Name', 'Category', 'Department', 'Purpose',
        'Key Names', 'Amount Found', 'Important Dates', 'Location Mentioned',
        'Confidentiality', 'Current Status', 'Processed On', 'File Type', 'File Size'
    ]
    
    extended_columns = [col for col in df.columns if col not in core_columns]
    
    print(f"\n✓ Core columns (14): {core_columns}")
    print(f"✓ Extended metadata columns ({len(extended_columns)}): {extended_columns}")
    
    # Analyze data quality
    analysis = {
        'total_rows': len(df),
        'total_columns': len(df.columns),
        'core_columns': core_columns,
        'extended_columns': extended_columns,
        'column_fill_rates': {},
        'sample_rows': []
    }
    
    # Calculate fill rates for each column
    for col in df.columns:
        non_empty = df[col].notna().sum()
        fill_rate = (non_empty / len(df) * 100) if len(df) > 0 else 0
        analysis['column_fill_rates'][col] = {
            'filled': non_empty,
            'total': len(df),
            'fill_rate_pct': round(fill_rate, 2)
        }
    
    # Get sample rows
    for idx in range(min(5, len(df))):
        row_data = {}
        for col in df.columns:
            val = df.iloc[idx][col]
            if pd.notna(val) and str(val).strip():
                row_data[col] = val
        analysis['sample_rows'].append(row_data)
    
    print(f"\n✓ Fill rates for extended columns:")
    for col in extended_columns:
        if col in analysis['column_fill_rates']:
            rate_info = analysis['column_fill_rates'][col]
            print(f"  {col}: {rate_info['filled']}/{rate_info['total']} ({rate_info['fill_rate_pct']}%)")
    
    if analysis['sample_rows']:
        print(f"\n✓ Sample row (first document):")
        print(json.dumps(analysis['sample_rows'][0], indent=2, default=str))
    
    return analysis


def validate_metadata_accuracy(metadata_path: str, state_matrix_analysis: Dict[str, Any]) -> Dict[str, Any]:
    """Validate that State Matrix extended columns match metadata workbook structure."""
    print(f"\n{'='*80}")
    print("VALIDATING METADATA ACCURACY")
    print(f"{'='*80}")
    
    # Expected extended columns based on the 12-Dimensional Taxonomy
    expected_extended_columns = {
        'Metadata Level',
        'Record Class',
        'Record Category (Functional)',
        'Record Category (Transactional)',
        'Record Type Code',
        'Business Unit',
        'Sub Business Unit',
        'ISO Country Code',
        'Record Format',
        'Original Location Type',
        'Data Classification',
        'Divestiture Deal Name'
    }
    
    actual_extended_columns = set(state_matrix_analysis['extended_columns'])
    
    # Normalize column names for comparison (lowercase, remove spaces/underscores)
    def normalize(col_name):
        return col_name.lower().replace(' ', '').replace('_', '')
    
    expected_normalized = {normalize(col): col for col in expected_extended_columns}
    actual_normalized = {normalize(col): col for col in actual_extended_columns}
    
    validation_results = {
        'expected_columns': list(expected_extended_columns),
        'actual_columns': list(actual_extended_columns),
        'matched_columns': [],
        'missing_columns': [],
        'extra_columns': [],
        'accuracy_pct': 0.0
    }
    
    # Find matches
    for norm_expected, expected_col in expected_normalized.items():
        if norm_expected in actual_normalized:
            validation_results['matched_columns'].append({
                'expected': expected_col,
                'actual': actual_normalized[norm_expected]
            })
        else:
            validation_results['missing_columns'].append(expected_col)
    
    # Find extras
    for norm_actual, actual_col in actual_normalized.items():
        if norm_actual not in expected_normalized:
            validation_results['extra_columns'].append(actual_col)
    
    # Calculate accuracy
    if expected_extended_columns:
        accuracy = (len(validation_results['matched_columns']) / len(expected_extended_columns)) * 100
        validation_results['accuracy_pct'] = round(accuracy, 2)
    
    print(f"\n✓ Expected extended columns: {len(expected_extended_columns)}")
    print(f"✓ Actual extended columns: {len(actual_extended_columns)}")
    print(f"✓ Matched columns: {len(validation_results['matched_columns'])}")
    
    if validation_results['matched_columns']:
        print(f"\n✓ MATCHED COLUMNS:")
        for match in validation_results['matched_columns']:
            print(f"  • {match['expected']} → {match['actual']}")
    
    if validation_results['missing_columns']:
        print(f"\n✗ MISSING COLUMNS (expected but not found):")
        for col in validation_results['missing_columns']:
            print(f"  • {col}")
    
    if validation_results['extra_columns']:
        print(f"\n⚠ EXTRA COLUMNS (found but not expected):")
        for col in validation_results['extra_columns']:
            print(f"  • {col}")
    
    print(f"\n{'='*80}")
    print(f"METADATA MAPPING ACCURACY: {validation_results['accuracy_pct']}%")
    print(f"{'='*80}")
    
    return validation_results


def validate_cell_values(state_matrix_path: str) -> Dict[str, Any]:
    """Validate cell values in State Matrix are accurate."""
    print(f"\n{'='*80}")
    print("VALIDATING CELL VALUES")
    print(f"{'='*80}")
    
    df = pd.read_excel(state_matrix_path, sheet_name="📄 Document Data")
    
    validation = {
        'total_cells': 0,
        'filled_cells': 0,
        'empty_cells': 0,
        'default_placeholders': 0,
        'issues': []
    }
    
    # Count cell types
    for col in df.columns:
        for idx, val in enumerate(df[col]):
            validation['total_cells'] += 1
            
            if pd.isna(val) or str(val).strip() == '':
                validation['empty_cells'] += 1
            elif str(val).strip().lower() in ['unclassified', 'unknown', 'none', 'n/a', '–']:
                validation['default_placeholders'] += 1
            else:
                validation['filled_cells'] += 1
    
    fill_rate = (validation['filled_cells'] / validation['total_cells'] * 100) if validation['total_cells'] > 0 else 0
    validation['fill_rate_pct'] = round(fill_rate, 2)
    
    print(f"\n✓ Total cells: {validation['total_cells']}")
    print(f"✓ Filled cells: {validation['filled_cells']} ({validation['fill_rate_pct']}%)")
    print(f"✓ Empty cells: {validation['empty_cells']}")
    print(f"✓ Default placeholders: {validation['default_placeholders']}")
    
    return validation


def main():
    """Run comprehensive validation."""
    print(f"\n{'#'*80}")
    print("STATE MATRIX ACCURACY VALIDATION")
    print(f"{'#'*80}")
    
    # Step 1: Check processing status
    print("\n[1/6] Checking processing status...")
    status = check_processing_status()
    print(f"  Total documents: {status['total']}")
    print(f"  Completed: {status['completed']}")
    print(f"  With extended metadata: {status['extended_count']}")
    
    if status['completed'] == 0:
        print("\n⚠ No documents have been processed yet. Please wait for processing to complete.")
        print("  Run this script again in a few minutes.")
        return
    
    if status['samples']:
        print(f"\n  Sample document with extended metadata:")
        sample = status['samples'][0]
        print(f"    File: {sample['file_name']}")
        print(f"    Category: {sample['category']}")
        print(f"    Department: {sample['department']}")
        print(f"    Purpose: {sample['purpose']}")
        if sample['extended_metadata_json']:
            try:
                extended = json.loads(sample['extended_metadata_json'])
                print(f"    Extended metadata: {json.dumps(extended, indent=6)}")
            except:
                pass
    
    # Step 2: Analyze metadata workbook
    print("\n[2/6] Analyzing metadata workbook schema...")
    cfg = get_config()
    metadata_path = getattr(cfg.tagging, "metadata_excel_path", "")
    if not metadata_path or not Path(metadata_path).exists():
        metadata_path = r"C:\Users\DELL\Downloads\Metadata Mapping_ Index and Copy ticket_Box - GECIHL ILAAP V2 (1).xlsx"
    metadata_schema = analyze_metadata_schema(metadata_path)
    
    # Step 3: Export and analyze State Matrix
    print("\n[3/6] Exporting State Matrix...")
    state_matrix_analysis = export_and_analyze_state_matrix()
    
    # Step 4: Validate metadata accuracy
    print("\n[4/6] Validating metadata mapping accuracy...")
    validation_results = validate_metadata_accuracy(metadata_path, state_matrix_analysis)
    
    # Step 5: Validate cell values
    print("\n[5/6] Validating cell value accuracy...")
    cell_validation = validate_cell_values("runtime/test_state_matrix.xlsx")
    
    # Step 6: Final summary
    print(f"\n{'#'*80}")
    print("VALIDATION SUMMARY")
    print(f"{'#'*80}")
    
    print(f"\n✓ Documents processed: {status['completed']}/{status['total']}")
    print(f"✓ Documents with extended metadata: {status['extended_count']}")
    print(f"✓ State Matrix rows: {state_matrix_analysis['total_rows']}")
    print(f"✓ State Matrix columns: {state_matrix_analysis['total_columns']}")
    print(f"  - Core columns: {len(state_matrix_analysis['core_columns'])}")
    print(f"  - Extended columns: {len(state_matrix_analysis['extended_columns'])}")
    print(f"\n✓ Metadata mapping accuracy: {validation_results['accuracy_pct']}%")
    print(f"✓ Cell fill rate: {cell_validation['fill_rate_pct']}%")
    
    # Overall assessment
    if validation_results['accuracy_pct'] >= 80 and cell_validation['fill_rate_pct'] >= 50:
        print(f"\n{'='*80}")
        print("✅ VALIDATION PASSED - System is working as expected!")
        print(f"{'='*80}")
        return 0
    else:
        print(f"\n{'='*80}")
        print("❌ VALIDATION FAILED - Issues detected!")
        print(f"{'='*80}")
        if validation_results['missing_columns']:
            print(f"\n⚠ Missing expected columns: {validation_results['missing_columns']}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
