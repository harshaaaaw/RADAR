"""Add missing categories to master_taxonomy.xlsx and enrich department keywords."""
import openpyxl
from pathlib import Path

TAXO_PATH = Path("runtime/taxonomy/master_taxonomy.xlsx")
wb = openpyxl.load_workbook(TAXO_PATH)

# ── Add to "category" sheet ──────────────────────────────────────────────
ws = wb["category"]

# Find headers
headers = [cell.value for cell in ws[1]]
print(f"Category headers: {headers}")

# Find the column indices
label_col = headers.index("label") + 1
aliases_col = headers.index("aliases") + 1
keywords_col = headers.index("keywords") + 1
active_col = headers.index("active") + 1
priority_col = headers.index("priority") + 1

next_row = ws.max_row + 1

# Add Offer Letter
ws.cell(row=next_row, column=label_col, value="Offer Letter")
ws.cell(row=next_row, column=aliases_col, value="appointment letter, employment letter, job offer, offer of employment")
ws.cell(row=next_row, column=keywords_col, value="offer, appointment, compensation, salary, joining date, designation, employment, position, annual package, benefits, probation, notice period")
ws.cell(row=next_row, column=active_col, value="TRUE")
ws.cell(row=next_row, column=priority_col, value=10)
print(f"Added Offer Letter at row {next_row}")

next_row += 1

# Add Meeting Minutes 
ws.cell(row=next_row, column=label_col, value="Meeting Minutes")
ws.cell(row=next_row, column=aliases_col, value="minutes of meeting, mom, board minutes, meeting notes, meeting record")
ws.cell(row=next_row, column=keywords_col, value="minutes, attendees, agenda, resolution, quorum, chairperson, board, meeting, proceedings, action items, voted, approved")
ws.cell(row=next_row, column=active_col, value="TRUE")
ws.cell(row=next_row, column=priority_col, value=9)
print(f"Added Meeting Minutes at row {next_row}")

# ── Enrich department keywords in "department" sheet ─────────────────────
ws_dept = wb["department"]
dept_headers = [cell.value for cell in ws_dept[1]]
dept_label_col = dept_headers.index("label") + 1
dept_kw_col = dept_headers.index("keywords") + 1
dept_aliases_col = dept_headers.index("aliases") + 1

for row_idx in range(2, ws_dept.max_row + 1):
    label = ws_dept.cell(row=row_idx, column=dept_label_col).value
    if not label:
        continue
    
    current_kw = ws_dept.cell(row=row_idx, column=dept_kw_col).value or ""
    current_aliases = ws_dept.cell(row=row_idx, column=dept_aliases_col).value or ""
    
    # Enrich Finance keywords
    if label == "Finance":
        if "chief financial officer" not in current_kw.lower():
            new_kw = current_kw + ", chief financial officer, cfo, accounts payable, accounts receivable, financial statement, treasury, taxation, income tax, gst return, debit, credit, bank, fiscal"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Finance keywords")
    
    # Enrich Legal keywords
    elif label == "Legal":
        if "service agreement" not in current_kw.lower():
            new_kw = current_kw + ", service agreement, terms and conditions, indemnity, jurisdiction, governing law, legal counsel, arbitration, parties, terminate, confidentiality, intellectual property"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Legal keywords")
    
    # Enrich Security keywords
    elif label == "Security":
        if "security incident" not in current_kw.lower():
            new_kw = current_kw + ", security incident, ciso, soc analyst, unauthorized access, brute force, credential stuffing, firewall, intrusion, penetration test, vulnerability scan, security operations center, remediation"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Security keywords")
    
    # Enrich Engineering keywords  
    elif label == "Engineering":
        if "devops" not in current_kw.lower():
            new_kw = current_kw + ", devops, infrastructure, latency, uptime, api, microservice, deployment pipeline, performance, system architecture, ec2, cloud, availability, sre"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Engineering keywords")
    
    # Enrich IT keywords
    elif label == "IT":
        if "laptop" not in current_kw.lower():
            new_kw = current_kw + ", laptop, desktop, hardware, software license, it procurement, helpdesk ticket, it manager, monitor, keyboard, workstation"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched IT keywords")
    
    # Enrich HR keywords
    elif label == "HR":
        if "offer letter" not in current_kw.lower():
            new_kw = current_kw + ", offer letter, appointment letter, compensation, salary, designation, joining date, probation, notice period, annual package, benefits"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched HR keywords")

    # Enrich Executive keywords
    elif label == "Executive":
        if "board meeting" not in current_kw.lower():
            new_kw = current_kw + ", board meeting, minutes, resolution, quorum, chairperson, board of directors, coo, ceo, managing director"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Executive keywords")

    # Enrich Compliance keywords
    elif label == "Compliance":
        if "data protection" not in current_kw.lower():
            new_kw = current_kw + ", data protection, gdpr, privacy, dpa, personal data, data subject, data breach notification, retention policy, right to erasure"
            ws_dept.cell(row=row_idx, column=dept_kw_col, value=new_kw)
            print(f"Enriched Compliance keywords")

wb.save(TAXO_PATH)
print(f"\nSaved taxonomy to {TAXO_PATH}")
