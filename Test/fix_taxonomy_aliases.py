"""Remove overly-short/generic aliases from taxonomy that cause false positives."""
import openpyxl

TAXONOMY_PATH = r"runtime\taxonomy\master_taxonomy.xlsx"
wb = openpyxl.load_workbook(TAXONOMY_PATH)

# Fix aliases sheet: remove "inv" (too short, matches inv-XXXXX reference codes)
ws = wb['aliases']
headers = [cell.value for cell in ws[1]]
field_col = headers.index('field') + 1
label_col = headers.index('label') + 1  
aliases_col = headers.index('aliases') + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    field = row[field_col - 1].value
    label = row[label_col - 1].value
    aliases = row[aliases_col - 1].value or ''
    
    if field == 'category' and label == 'Invoice':
        # Remove "inv" and "inv no" — too short, match reference codes like INV-31318
        alias_list = [a.strip() for a in aliases.split(',') if a.strip()]
        filtered = [a for a in alias_list if a not in ('inv', 'inv no')]
        row[aliases_col - 1].value = ','.join(filtered)
        print(f"  Fixed Invoice aliases: {aliases} -> {row[aliases_col - 1].value}")

# Fix category sheet: remove "record" alias from Data Record (too generic)
ws = wb['category']
headers = [cell.value for cell in ws[1]]
label_col = headers.index('label') + 1
aliases_col = headers.index('aliases') + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    label = row[label_col - 1].value
    if label == 'Data Record':
        aliases = row[aliases_col - 1].value or ''
        alias_list = [a.strip() for a in aliases.split(',') if a.strip()]
        filtered = [a for a in alias_list if a != 'record']
        row[aliases_col - 1].value = ','.join(filtered)
        print(f"  Fixed Data Record aliases: {aliases} -> {row[aliases_col - 1].value}")

# Also fix the aliases sheet Data Record entry: remove "record" from category/Data Record 
ws = wb['aliases']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    field = row[field_col - 1].value
    label = row[label_col - 1].value
    aliases = row[aliases_col - 1].value or ''
    if field == 'category' and label == 'Budget':
        # Remove "fiscal" (too short/generic, matches in mixed content) if alone
        alias_list = [a.strip() for a in aliases.split(',') if a.strip()]
        filtered = [a for a in alias_list if a != 'fiscal']
        row[aliases_col - 1].value = ','.join(filtered)
        print(f"  Fixed Budget aliases: {aliases} -> {row[aliases_col - 1].value}")

wb.save(TAXONOMY_PATH)
print(f"\nSaved updated taxonomy")
