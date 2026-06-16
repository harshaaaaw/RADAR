"""Update master_taxonomy.xlsx to fix tagging issues:
1. Email category: Remove stop-word keywords 'from','to' — replace with specific phrases
2. Department: Add label names as aliases (engineering,finance,sales etc.)
3. Purpose: Expand keyword coverage, lower Record Keeping priority
4. Add Budget category to category sheet
"""
import openpyxl

TAXONOMY_PATH = r"runtime\taxonomy\master_taxonomy.xlsx"
wb = openpyxl.load_workbook(TAXONOMY_PATH)

# ========================================
# FIX 1: Category sheet — Fix Email keywords + add Budget
# ========================================
ws = wb['category']
headers = [cell.value for cell in ws[1]]
label_col = headers.index('label') + 1
aliases_col = headers.index('aliases') + 1
keywords_col = headers.index('keywords') + 1
active_col = headers.index('active') + 1
priority_col = headers.index('priority') + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    label = row[label_col - 1].value
    if label == 'Email':
        # Remove stop-word keywords, replace with specific email phrases
        row[keywords_col - 1].value = 'email thread,email chain,forwarded message,reply to,inbox,subject,dear,regards,bcc'
        row[aliases_col - 1].value = 'mail,thread,communication,email correspondence,email message'
        print(f"  Fixed Email keywords: {row[keywords_col - 1].value}")

# Add Budget category
next_row = ws.max_row + 1
ws.cell(row=next_row, column=label_col, value='Budget')
ws.cell(row=next_row, column=aliases_col, value='budget report,budget plan,fiscal plan,financial plan')
ws.cell(row=next_row, column=keywords_col, value='budget,allocation,forecast,fiscal year,expenditure,planned spend,revenue projection')
ws.cell(row=next_row, column=active_col, value='true')
ws.cell(row=next_row, column=priority_col, value=8)
print(f"  Added Budget category row")

# Add Data Record / Data Entry category
next_row = ws.max_row + 1
ws.cell(row=next_row, column=label_col, value='Data Record')
ws.cell(row=next_row, column=aliases_col, value='data entry,data record,record,data file')
ws.cell(row=next_row, column=keywords_col, value='data,record,entry,field,value,column,row,dataset')
ws.cell(row=next_row, column=active_col, value='true')
ws.cell(row=next_row, column=priority_col, value=7)
print(f"  Added Data Record category row")

# ========================================
# FIX 2: Department sheet — Add label names as aliases
# ========================================
ws = wb['department']
headers = [cell.value for cell in ws[1]]
label_col = headers.index('label') + 1
aliases_col = headers.index('aliases') + 1
keywords_col = headers.index('keywords') + 1

dept_alias_additions = {
    'Engineering': 'engineering,eng',
    'Finance': 'finance,financial',
    'IT': 'technology',
    'Sales': 'sales',
    'HR': 'hr',
    'Operations': 'operations',
    'Marketing': 'marketing',
    'Legal': 'legal',
    'Procurement': 'procurement',
    'Customer Support': 'customer support,support',
    'Security': 'security',
    'Compliance': 'compliance',
    'Administration': 'administration,admin',
    'Executive': 'executive',
}

# Also expand department keywords for better matching
dept_keyword_additions = {
    'Finance': ',budget,allocation,fiscal,cost,profit,finance',
    'Engineering': ',engineering,design,technical,specification,prototype',
    'Sales': ',sales,revenue,target,commission,quota,customer',
    'IT': ',it support,system,database,cloud,software,hardware',
    'HR': ',human resources,personnel,staff,workforce,benefits',
}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    label = row[label_col - 1].value
    if label in dept_alias_additions:
        current_aliases = row[aliases_col - 1].value or ''
        new_aliases = dept_alias_additions[label]
        # Only add if not already present
        existing = set(a.strip().lower() for a in current_aliases.split(',') if a.strip())
        to_add = [a.strip() for a in new_aliases.split(',') if a.strip().lower() not in existing]
        if to_add:
            updated = current_aliases + ',' + ','.join(to_add) if current_aliases else ','.join(to_add)
            row[aliases_col - 1].value = updated
            print(f"  Department '{label}': added aliases {to_add}")
    
    if label in dept_keyword_additions:
        current_kw = row[keywords_col - 1].value or ''
        existing_kw = set(k.strip().lower() for k in current_kw.split(',') if k.strip())
        new_kws = [k.strip() for k in dept_keyword_additions[label].split(',') if k.strip().lower() not in existing_kw]
        if new_kws:
            row[keywords_col - 1].value = current_kw + ',' + ','.join(new_kws)
            print(f"  Department '{label}': added keywords {new_kws}")

# ========================================
# FIX 3: Purpose sheet — expand keywords, lower Record Keeping priority
# ========================================
ws = wb['purpose']
headers = [cell.value for cell in ws[1]]
label_col = headers.index('label') + 1
aliases_col = headers.index('aliases') + 1
keywords_col = headers.index('keywords') + 1
priority_col = headers.index('priority') + 1

purpose_keyword_additions = {
    'Payment Processing': ',invoice,amount,payment,dues,billing',
    'Reporting': ',report,analysis,summary,dashboard,metrics,overview',
}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    label = row[label_col - 1].value
    if label == 'Record Keeping':
        # Lower priority so it only wins when nothing else matches
        row[priority_col - 1].value = 3
        print(f"  Lowered Record Keeping priority to 3")
    
    if label in purpose_keyword_additions:
        current_kw = row[keywords_col - 1].value or ''
        existing_kw = set(k.strip().lower() for k in current_kw.split(',') if k.strip())
        new_kws = [k.strip() for k in purpose_keyword_additions[label].split(',') if k.strip().lower() not in existing_kw]
        if new_kws:
            row[keywords_col - 1].value = current_kw + ',' + ','.join(new_kws)
            print(f"  Purpose '{label}': added keywords {new_kws}")

# Add a general purpose "Documentation" for budget/report type docs
next_row = ws.max_row + 1
ws.cell(row=next_row, column=label_col, value='Documentation')
ws.cell(row=next_row, column=aliases_col, value='documentation,document,doc')
ws.cell(row=next_row, column=keywords_col, value='document,documentation,reference,information,details,specification')
ws.cell(row=next_row, column=headers.index('active') + 1, value='true')
ws.cell(row=next_row, column=priority_col, value=5)
print(f"  Added Documentation purpose row")

# ========================================
# FIX 4: Aliases sheet — add department label names
# ========================================
ws = wb['aliases']
headers = [cell.value for cell in ws[1]]
field_col = headers.index('field') + 1
label_col_a = headers.index('label') + 1
aliases_col_a = headers.index('aliases') + 1

# Add Engineering department alias for better matching
existing_aliases = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] and row[1]:
        existing_aliases.add(f"{row[0]}:{row[1]}")

new_aliases = [
    ('department', 'Engineering', 'engineering,eng,tech engineering'),
    ('department', 'Sales', 'sales,business development,biz dev'),
    ('department', 'Operations', 'operations,ops,operational'),
    ('department', 'Marketing', 'marketing,digital marketing,brand'),
    ('category', 'Budget', 'budget,budget report,budget plan,fiscal,financial budget'),
]

for field, label, aliases in new_aliases:
    key = f"{field}:{label}"
    if key not in existing_aliases:
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=field_col, value=field)
        ws.cell(row=next_row, column=label_col_a, value=label)
        ws.cell(row=next_row, column=aliases_col_a, value=aliases)
        print(f"  Added alias row: {field}/{label}: {aliases}")

wb.save(TAXONOMY_PATH)
print(f"\nSaved updated taxonomy to: {TAXONOMY_PATH}")
