"""Dump taxonomy Excel sheets to understand current keywords/aliases/defaults."""
import openpyxl

wb = openpyxl.load_workbook(r"runtime\taxonomy\master_taxonomy.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"{'='*80}")
    print(f"SHEET: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*80}")
    
    headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
    print(f"Headers: {headers}")
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [str(v) if v is not None else '' for v in row]
        if any(v for v in vals):
            print(f"  {' | '.join(vals)}")
    print()
