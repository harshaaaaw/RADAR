"""
Metadata workbook manager for metadata-first tagging.
"""

from __future__ import annotations

import hashlib
import json
import re
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from core.config_manager import get_config
from core.logging_manager import get_logger

from .tagging_models import TaggingRequest

logger = get_logger("tagging.metadata")
_TOKEN_RE = re.compile(r"[a-z0-9]{3,}")


class Sheet3ValidValuesRegistry:
    """
    Parses Sheet 3 (3.List of Valid Values) from the metadata Excel workbook.
    Builds a per-field allowed-values set used for strict constraint enforcement.
    
    Sheet 3 has these columns (row 4 = header, rows 5+ = data):
      Col 1:  Metadata Level Code          → allowed: {File, Folder}
      Col 2:  Record Class Name            → allowed: {Functional, Transactional, Customer, Undefined}
      Col 3:  Record Category Name (Functional)  → 22 values
      Col 4:  Record Category Name (Transactional/Customer) → 30 values
      Col 5:  Record Type Code             → ~90 codes
      Col 7:  Business Unit Name           → 5 values
      Col 8:  Sub Business Unit Name       → 32 values
      Col 9:  ISO Country Code             → ~250 codes
      Col 11: Record Format Name           → {Electronic, Physical}
      Col 12: Original Record Location Type Name → 6 values
      Col 13: Data Classification Name     → 5 values
      Col 14: Divestiture Deal Name        → 51 values
    """
    
    SHEET3_NAME = "3.List of Valid Values"
    HEADER_ROW = 4   # 1-indexed (row 4 in Excel)
    DATA_START_ROW = 5
    
    # Maps sheet column index (1-based) → registry field key
    COLUMN_FIELD_MAP = {
        1:  "metadata_level_code",
        2:  "record_class_name",
        3:  "record_category_name_functional",
        4:  "record_category_name_transactional",
        5:  "record_type_code",
        7:  "business_unit_name",
        8:  "sub_business_unit_name",
        9:  "iso_country_code",
        11: "record_format_name",
        12: "original_record_location_type_name",
        13: "data_classification_name",
        14: "divestiture_deal_name",
    }
    
    # Maps registry field key → tagging output field
    FIELD_TO_TAG = {
        "record_class_name":                   "category",        # secondary signal
        "record_category_name_functional":     "category",        # primary
        "record_category_name_transactional":  "category",        # secondary
        "business_unit_name":                  "department",      # primary
        "data_classification_name":            "confidentiality", # direct
    }

    @classmethod
    def load_from_workbook(cls, path: str | Path) -> Dict[str, Set[str]]:
        import openpyxl
        
        path = Path(path)
        if not path.exists():
            logger.warning("Workbook path %s does not exist. Cannot load Sheet 3.", path)
            return {}
            
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            
            sheet = None
            for name in wb.sheetnames:
                if name.strip().lower() == cls.SHEET3_NAME.lower():
                    sheet = wb[name]
                    break
            
            if sheet is None:
                for name in wb.sheetnames:
                    if "valid" in name.lower() and "values" in name.lower():
                        sheet = wb[name]
                        break
            
            if sheet is None:
                logger.warning("Sheet %s not found in workbook %s", cls.SHEET3_NAME, path)
                return {}
            
            registry: Dict[str, Set[str]] = {key: set() for key in cls.COLUMN_FIELD_MAP.values()}
            
            for row in sheet.iter_rows(min_row=cls.DATA_START_ROW, values_only=True):
                for col_idx, field_key in cls.COLUMN_FIELD_MAP.items():
                    idx_0 = col_idx - 1
                    if idx_0 < len(row):
                        val = row[idx_0]
                        if val is not None:
                            val_str = str(val).strip()
                            if val_str and val_str.lower() not in {"na", "n/a", "none", "null", "unknown", "not applicable", "-"}:
                                registry[field_key].add(val_str)
            
            for k, v in registry.items():
                logger.debug("Loaded %d valid values for field %s", len(v), k)
            logger.info("Successfully loaded Sheet 3 valid values registry from %s (fields: %d)", path, len(registry))
            return registry
            
        except Exception as e:
            logger.exception("Error loading Sheet 3 registry from %s: %r", path, e)
            return {}
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @classmethod
    def find_best_allowed_value(cls, candidate: str, allowed_set: Set[str]) -> Optional[str]:
        if not candidate or not allowed_set:
            return None
            
        candidate_norm = _norm(candidate)
        if not candidate_norm:
            return None
            
        # 1. Exact match (case-insensitive, normalized)
        for allowed in allowed_set:
            if _norm(allowed) == candidate_norm:
                return allowed
                
        # 2. Partial match (candidate is substring of an allowed value or vice versa)
        if len(candidate_norm) >= 4:
            for allowed in allowed_set:
                allowed_norm = _norm(allowed)
                if candidate_norm in allowed_norm or allowed_norm in candidate_norm:
                    return allowed
                    
        # 3. Token overlap (tokenize candidate, find allowed with max word overlap)
        candidate_tokens = set(_TOKEN_RE.findall(candidate_norm))
        if candidate_tokens:
            best_match = None
            max_overlap = 0
            for allowed in allowed_set:
                allowed_norm = _norm(allowed)
                allowed_tokens = set(_TOKEN_RE.findall(allowed_norm))
                overlap = len(candidate_tokens.intersection(allowed_tokens))
                if overlap > max_overlap:
                    max_overlap = overlap
                    best_match = allowed
            if max_overlap > 0:
                return best_match
                
        # 4. Synonym Fallback Mapping for Disjoint Vocabularies
        # Normalize the allowed set for robust matching
        allowed_norms = {str(x).strip().lower() for x in allowed_set}
        
        # Check if this allowed_set is the Department (Business Unit) set
        is_dept = (
            "gecc hq" in allowed_norms or 
            "ge capital international" in allowed_norms or 
            "treasury" in allowed_norms or
            any("business unit" in x or "business unite" in x or "gecchq" in x for x in allowed_norms)
        )
        # Check if this allowed_set is the Category (Functional) set
        is_cat = (
            "finance & accounting" in allowed_norms or 
            "common - all departments" in allowed_norms or
            any("functional" in x or "finance & accounting" in x for x in allowed_norms)
        )
        # Check if this allowed_set is the Confidentiality set
        is_conf = (
            "ge confidential" in allowed_norms or 
            "ge internal" in allowed_norms or
            any("data classification" in x or "ge confidential" in x or "classification name" in x for x in allowed_norms)
        )

        if is_dept:
            cand_lower = candidate.strip().lower()
            if any(w in cand_lower for w in ["finance", "accounting", "treasury", "fin"]):
                return "Treasury"
            if any(w in cand_lower for w in ["hr", "human resource", "personnel", "admin", "it", "engineering", "r&d", "general", "unclassified"]):
                return "GECC HQ"
            if any(w in cand_lower for w in ["operations", "sales", "marketing", "customer service", "ops", "support"]):
                return "GE Capital International"
            if "america" in cand_lower:
                return "Americas"
            if "estate" in cand_lower or "property" in cand_lower:
                return "Real Estate"
            return "GECC HQ"

        if is_cat:
            cand_lower = candidate.strip().lower()
            if any(w in cand_lower for w in ["finance", "accounting", "invoice", "budget", "fin & acct"]):
                return "Finance & Accounting"
            if any(w in cand_lower for w in ["hr", "human resource", "people", "employee", "performance", "recruiting"]):
                return "Human Resources"
            if any(w in cand_lower for w in ["it", "technology", "information technology", "tech", "software", "hardware"]):
                return "Information Technology"
            if any(w in cand_lower for w in ["legal", "law", "contract", "agreement", "nda"]):
                return "Legal"
            if any(w in cand_lower for w in ["compliance", "regulatory", "regulation"]):
                return "Regulatory Management & Compliance"
            if any(w in cand_lower for w in ["audit", "inspection"]):
                return "Audit"
            if any(w in cand_lower for w in ["tax", "levy"]):
                return "Tax"
            if any(w in cand_lower for w in ["risk", "loss", "credit risk"]):
                return "Risk Management"
            if any(w in cand_lower for w in ["facilities", "facility", "building", "office"]):
                return "Facilities Management"
            if any(w in cand_lower for w in ["marketing", "ad ", "advertising"]):
                return "Marketing"
            if any(w in cand_lower for w in ["sourcing", "procurement", "vendor"]):
                return "Sourcing"
            if any(w in cand_lower for w in ["general", "unclassified", "other", "misc", "common"]):
                return "Common - All Departments"
            return "Common - All Departments"

        if is_conf:
            cand_lower = candidate.strip().lower()
            if "spii" in cand_lower:
                return "GE Confidential with SPII"
            if "restricted" in cand_lower:
                return "GE Restricted"
            if "confidential" in cand_lower:
                return "GE Confidential"
            if "internal" in cand_lower:
                return "GE Internal"
            if "public" in cand_lower:
                return "Public"
            return "GE Internal"
            
        return None


def _norm(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _checksum(path: Path) -> str:
    try:
        return hashlib.sha1(path.read_bytes()).hexdigest()
    except Exception:
        return ""


@dataclass
class MetadataSnapshot:
    path: str
    checksum: str
    loaded_at: float
    source: str
    sheet_names: List[str]
    rows: List[Dict[str, Any]]
    detected_schema: Dict[str, Any] = None  # Dynamic schema info
    field_mappings: Dict[str, str] = None  # Column → core field mappings
    extended_columns: List[str] = None  # Non-core columns to preserve
    sheet3_allowed_values: Dict[str, Set[str]] = None  # Populated from Sheet 3


class MetadataManager:
    """Loads metadata workbook and resolves metadata tags for a tagging request."""

    SOURCE_PRIORITY = {
        "config": 10,
        "api": 20,
        "ui": 30,
        "cli": 40,
    }

    TAG_ALIASES = {
        "category": [
            "category",
            "document category",
            "doc category",
            "business category",
            "record category",
            "record category name",
            "document type",
            "type",
        ],
        "department": [
            "department",
            "dept",
            "domain",
            "business unit",
            "business unit name",
            "function",
            "owner department",
        ],
        "purpose": [
            "purpose",
            "intent",
            "use case",
            "business purpose",
            "business function",
            "business function name",
            "objective",
        ],
        "confidentiality": [
            "confidentiality",
            "classification",
            "data classification",
            "data classification name",
            "sensitivity",
            "security classification",
            "privacy",
        ],
    }

    # Extended metadata columns to always preserve (beyond core fields)
    EXTENDED_METADATA_PATTERNS = [
        r"metadata.*level.*code",
        r"record.*class.*name",
        r"record.*category.*name",
        r"sub.*business.*unit",
        r"business.*unit",
        r"record.*type.*code",
        r"record.*type.*description",
        r"record.*type.*desc",
        r"record.*class",
        r"commentary",
        r"business.*function.*name",
        r"document.*type.*name",
        r"iso.*country",
        r"country.*name",
        r"transaction.*number",
        r"transaction.*name",
        r"transaction.*type.*name",
        r"customer.*number",
        r"customer.*name",
        r"record.*format.*name",
        r"original.*record.*location.*type.*name",
        r"record.*source.*name",
        r"data.*classification.*name",
        r"divestiture.*deal.*name",
        r"creation.*date",
        r"last.*modified.*date",
        r"creator.*name",
        r"disposition",
        r"status",
        r"staging.*file.*path.*location",
        r"original.*file.*path.*name",
        r"effective.*time",
        r"preservation.*hold",
        r"tax.*related",
        r"divestiture",
        r"record.*format",
        r"original.*record.*location",
        r"record.*source",
        r"staging.*file.*path",
    ]

    EXCLUDED_EXTENDED_PATTERNS = [
        r"^non-metadata$",
        r"^metadata.*element.*name$",
        r"^mandatory.*optional.*indicator$",
        r"^mandatory.*metadata.*requirements$",
        r"^optional.*metadata.*elements$",
        r"^definition$",
        r"^allowable.*values",
        r"^metadata.*element.*format$",
        r"^data.*owner.*documentation",
        r"^irm.*checklist$",
    ]

    EXCLUDED_SHEET_PATTERNS = [
        r"metadata\s*definition",
        r"copy\s*ticket",
        r"valid\s*values",
    ]

    HEADER_ANCHORS = [
        "business unit",
        "sub business",
        "record type code",
        "record category",
        "record class",
        "metadata level code",
        "staging file path",
        "original file path",
        "document type",
        "customer number",
        "transaction number",
        "metadata element name",
    ]

    PLACEHOLDER_VALUES = {
        "na",
        "n/a",
        "none",
        "null",
        "unknown",
        "not available",
        "not_applicable",
        "-",
        "--",
    }

    JOIN_ALIASES = {
        "smart_id": ["smart id", "smart_id"],
        "file_key": ["file key", "file_key", "file hash", "file_hash"],
        "file_name": ["file name", "filename", "name", "document name", "doc name", "folder/app name"],
        "file_path": [
            "file path",
            "filepath",
            "path",
            "document path",
            "source path",
            "staging file path location",
            "original file path name",
        ],
        "doc_id": ["doc id", "document id", "doc_id", "id"],
    }

    def __init__(self) -> None:
        cfg = get_config()
        tagging_cfg = getattr(cfg, "tagging", None)
        self._default_path = str(getattr(tagging_cfg, "metadata_excel_path", "") or "")
        self._hot_reload_seconds = max(5, int(getattr(tagging_cfg, "hot_reload_seconds", 30) or 30))

        metadata_root = Path(cfg.paths.working_root) / "metadata"
        metadata_root.mkdir(parents=True, exist_ok=True)
        upload_dir = str(getattr(tagging_cfg, "metadata_upload_dir", "") or "")
        self.upload_dir = Path(upload_dir) if upload_dir else metadata_root / "uploads"
        self.upload_dir.mkdir(parents=True, exist_ok=True)

        self._active_descriptor_path = metadata_root / "active_metadata.json"
        self._snapshot: Optional[MetadataSnapshot] = None
        self._last_check = 0.0
        self._lock = threading.Lock()

    def set_active_metadata_source(self, path: str, source: str, *, force: bool = False) -> Tuple[bool, str]:
        src = _norm(source)
        if src not in self.SOURCE_PRIORITY:
            return False, f"Unsupported metadata source: {source}"

        file_path = Path(str(path or "")).expanduser().resolve()
        if not file_path.exists():
            return False, f"Metadata file does not exist: {file_path}"
        if file_path.suffix.lower() != ".xlsx":
            return False, "Metadata file must be an .xlsx workbook"

        current = self._read_active_descriptor()
        if current and not force:
            cur_source = _norm(current.get("source", "config"))
            if self.SOURCE_PRIORITY.get(cur_source, 0) > self.SOURCE_PRIORITY[src]:
                return False, f"Active metadata source '{cur_source}' has higher priority than '{src}'"

        descriptor = {
            "path": str(file_path),
            "source": src,
            "updated_at": time.time(),
            "checksum": _checksum(file_path),
        }
        self._active_descriptor_path.write_text(json.dumps(descriptor, indent=2), encoding="utf-8")
        with self._lock:
            self._snapshot = None
            self._last_check = 0.0
        logger.info("Metadata source set via %s: %s", src, file_path)
        return True, "Metadata source updated"

    def clear_active_metadata_source(self) -> None:
        try:
            descriptor = {
                "path": "",
                "source": "cli",
                "disabled": True,
                "updated_at": time.time(),
            }
            self._active_descriptor_path.write_text(json.dumps(descriptor, indent=2), encoding="utf-8")
        except Exception:
            pass
        with self._lock:
            self._snapshot = None
            self._last_check = 0.0

    def get_status(self) -> Dict[str, Any]:
        descriptor = self._resolve_descriptor()
        if not descriptor:
            return {
                "active": False,
                "mode": "spacy_only_mode",
                "source": "",
                "path": "",
                "message": "No metadata file configured",
            }

        path = Path(descriptor["path"])
        if not path.exists():
            return {
                "active": False,
                "mode": "spacy_only_mode",
                "source": descriptor.get("source", ""),
                "path": descriptor.get("path", ""),
                "message": "Configured metadata file does not exist",
            }

        return {
            "active": True,
            "mode": "metadata_mode",
            "source": descriptor.get("source", ""),
            "path": str(path),
            "checksum": descriptor.get("checksum", ""),
            "updated_at": descriptor.get("updated_at", 0.0),
            "message": "Metadata file active",
        }

    def ensure_loaded(self) -> Optional[MetadataSnapshot]:
        now = time.time()
        if self._snapshot and (now - self._last_check) < self._hot_reload_seconds:
            return self._snapshot

        with self._lock:
            if self._snapshot and (time.time() - self._last_check) < self._hot_reload_seconds:
                return self._snapshot

            self._last_check = time.time()
            descriptor = self._resolve_descriptor()
            if not descriptor:
                self._snapshot = None
                return None

            path = Path(descriptor["path"])
            if not path.exists():
                self._snapshot = None
                return None

            checksum = _checksum(path)
            if self._snapshot and self._snapshot.path == str(path) and self._snapshot.checksum == checksum:
                return self._snapshot

            self._snapshot = self._load_workbook(path=path, source=str(descriptor.get("source", "config") or "config"))
            return self._snapshot

    def get_sheet3_registry(self) -> Dict[str, Set[str]]:
        snapshot = self.ensure_loaded()
        if snapshot and snapshot.sheet3_allowed_values:
            return snapshot.sheet3_allowed_values
        return {}

    def resolve_tags(self, req: TaggingRequest) -> Dict[str, Any]:
        snapshot = self.ensure_loaded()
        if snapshot is None:
            return {
                "active": False,
                "mode": "spacy_only_mode",
                "source": "",
                "matched": False,
                "explicit": {},
                "derived": {},
                "extended": {},
                "reasons": ["metadata_unavailable"],
            }

        request_keys = self._build_request_keys(req)
        request_text = " ".join(
            [
                req.file_name or "",
                req.file_path or "",
                req.main_content or "",
                req.ocr_content or "",
                req.embedded_content or "",
            ]
        )
        request_tokens = set(_TOKEN_RE.findall(_norm(request_text)))
        best_row = None
        best_score = -1

        for row in snapshot.rows:
            score = self._match_score(request_keys, row.get("join_values", {}))
            if score > best_score:
                best_score = score
                best_row = row

        if best_row is None or best_score <= 0:
            # If workbook effectively provides a single metadata row, treat it as
            # folder-level metadata and apply deterministically.
            if len(snapshot.rows) == 1:
                only_row = snapshot.rows[0]
                return {
                    "active": True,
                    "mode": "metadata_mode",
                    "source": snapshot.source,
                    "matched": True,
                    "explicit": dict(only_row.get("explicit", {})),
                    "derived": dict(only_row.get("derived", {})),
                    "extended": dict(only_row.get("extended", {})),
                    "sheet": only_row.get("sheet", ""),
                    "row_index": only_row.get("row_index", 0),
                    "reasons": ["metadata_single_row_fallback"],
                }

            # Fallback for mapping-style metadata workbooks where file IDs may be absent:
            # use token overlap between request content and metadata row content.
            semantic_best_row = None
            semantic_best_score = 0
            for row in snapshot.rows:
                row_tokens = set(row.get("row_tokens", []))
                if not row_tokens:
                    continue
                overlap = len(request_tokens.intersection(row_tokens))
                if overlap > semantic_best_score:
                    semantic_best_score = overlap
                    semantic_best_row = row

            if semantic_best_row is not None and semantic_best_score >= 3:
                return {
                    "active": True,
                    "mode": "metadata_mode",
                    "source": snapshot.source,
                    "matched": True,
                    "explicit": dict(semantic_best_row.get("explicit", {})),
                    "derived": dict(semantic_best_row.get("derived", {})),
                    "extended": dict(semantic_best_row.get("extended", {})),
                    "sheet": semantic_best_row.get("sheet", ""),
                    "row_index": semantic_best_row.get("row_index", 0),
                    "reasons": [f"metadata_semantic_match_score:{semantic_best_score}"],
                }

            return {
                "active": True,
                "mode": "metadata_mode",
                "source": snapshot.source,
                "matched": False,
                "explicit": {},
                "derived": {},
                "extended": {},
                "reasons": ["metadata_no_row_match"],
            }

        return {
            "active": True,
            "mode": "metadata_mode",
            "source": snapshot.source,
            "matched": True,
            "explicit": dict(best_row.get("explicit", {})),
            "derived": dict(best_row.get("derived", {})),
            "extended": dict(best_row.get("extended", {})),
            "sheet": best_row.get("sheet", ""),
            "row_index": best_row.get("row_index", 0),
            "reasons": [f"metadata_row_match_score:{best_score}"],
        }

    def _resolve_descriptor(self) -> Optional[Dict[str, Any]]:
        active = self._read_active_descriptor()
        if active:
            if active.get("disabled") or not active.get("path"):
                # Disabled descriptor — fall through to config default below
                pass
            elif Path(str(active.get("path", ""))).exists():
                return active
            else:
                return None

        if self._default_path:
            cfg_path = Path(self._default_path).expanduser().resolve()
            if cfg_path.exists() and cfg_path.suffix.lower() == ".xlsx":
                return {
                    "path": str(cfg_path),
                    "source": "config",
                    "updated_at": 0.0,
                    "checksum": _checksum(cfg_path),
                }
        return None

    def _read_active_descriptor(self) -> Optional[Dict[str, Any]]:
        if not self._active_descriptor_path.exists():
            return None
        try:
            payload = json.loads(self._active_descriptor_path.read_text(encoding="utf-8"))
            if not isinstance(payload, dict):
                return None
            return payload
        except Exception:
            return None

    def _load_workbook(self, *, path: Path, source: str) -> MetadataSnapshot:
        import pandas as pd

        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
        prepared_sheets = {sheet_name: self._prepare_sheet_frame(frame) for sheet_name, frame in sheets.items()}
        rows: List[Dict[str, Any]] = []
        
        detected_schema = self._detect_schema(prepared_sheets)
        field_mappings = detected_schema.get("field_mappings", {})
        extended_columns = detected_schema.get("extended_columns", [])

        for sheet_name, frame in prepared_sheets.items():
            if self._is_excluded_sheet(sheet_name):
                continue
            normalized = frame.rename(columns={c: _norm(c) for c in frame.columns}).fillna("")
            for idx, raw in normalized.iterrows():
                values = {
                    str(k): str(v).strip()
                    for k, v in raw.items()
                    if str(v).strip() and not self._is_placeholder_value(str(v))
                }
                if not values:
                    continue

                explicit = self._extract_explicit_tags(values)
                derived = self._derive_tags(values)
                join_values = self._extract_join_values(values)
                extended = self._extract_extended_metadata(values, extended_columns)

                rows.append(
                    {
                        "sheet": sheet_name,
                        "row_index": int(idx) + 2,
                        "values": values,
                        "join_values": join_values,
                        "explicit": explicit,
                        "derived": derived,
                        "extended": extended,
                        "row_tokens": list(set(_TOKEN_RE.findall(_norm(" ".join(values.values()))))),
                    }
                )

        # Load Sheet 3 registry
        sheet3_registry = Sheet3ValidValuesRegistry.load_from_workbook(path)

        logger.info(
            "Loaded metadata workbook: %s (rows=%s, sheets=%s, extended_cols=%s)",
            path, len(rows), len(prepared_sheets.keys()), len(extended_columns)
        )
        return MetadataSnapshot(
            path=str(path),
            checksum=_checksum(path),
            loaded_at=time.time(),
            source=source,
            sheet_names=list(prepared_sheets.keys()),
            rows=rows,
            detected_schema=detected_schema,
            field_mappings=field_mappings,
            extended_columns=extended_columns,
            sheet3_allowed_values=sheet3_registry,
        )

    def _detect_schema(self, sheets: Dict[str, Any]) -> Dict[str, Any]:
        """Detect schema by analyzing all columns across sheets."""
        all_columns = set()
        for frame in sheets.values():
            all_columns.update(_norm(c) for c in frame.columns)
        
        # Identify which columns map to core fields
        field_mappings = {}
        for col in all_columns:
            for field, aliases in self.TAG_ALIASES.items():
                if col in [_norm(a) for a in aliases]:
                    field_mappings[col] = field
                    break
        
        # Identify extended metadata columns (not core, not join, not unnamed)
        core_mapped = set(field_mappings.keys())
        join_mapped = set()
        for aliases in self.JOIN_ALIASES.values():
            join_mapped.update(_norm(a) for a in aliases)
        
        extended_columns = []
        for col in sorted(all_columns):
            # Skip if already mapped to core or join fields
            if col in core_mapped or col in join_mapped:
                continue
            # Skip unnamed/empty columns
            if not col or col.startswith("unnamed"):
                continue
            # Skip known noisy template-definition columns
            if any(re.search(pattern, col, re.IGNORECASE) for pattern in self.EXCLUDED_EXTENDED_PATTERNS):
                continue
            # Check if matches extended metadata patterns
            is_extended = False
            for pattern in self.EXTENDED_METADATA_PATTERNS:
                if re.search(pattern, col, re.IGNORECASE):
                    is_extended = True
                    break
            if is_extended:
                extended_columns.append(col)
        
        return {
            "field_mappings": field_mappings,
            "extended_columns": extended_columns,
            "all_columns": list(all_columns),
        }

    def _extract_extended_metadata(self, values: Dict[str, str], extended_columns: List[str]) -> Dict[str, str]:
        """Extract extended metadata fields from row values."""
        extended = {}
        for col in extended_columns:
            value = values.get(col, "")
            if value and not self._is_placeholder_value(value):
                # Use original column name (from normalized) as key
                extended[col] = value
        return extended

    def _extract_explicit_tags(self, values: Dict[str, str]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for field, aliases in self.TAG_ALIASES.items():
            for alias in aliases:
                key = _norm(alias)
                value = values.get(key, "")
                if value and not self._is_placeholder_value(value):
                    out[field] = value
                    break
        return out

    def _is_placeholder_value(self, value: str) -> bool:
        return _norm(value) in self.PLACEHOLDER_VALUES

    def _is_excluded_sheet(self, sheet_name: str) -> bool:
        name = _norm(sheet_name)
        return any(re.search(pattern, name, re.IGNORECASE) for pattern in self.EXCLUDED_SHEET_PATTERNS)

    def _prepare_sheet_frame(self, frame: Any) -> Any:
        """Normalize sheet layout by promoting in-sheet header rows when needed."""
        safe = frame.fillna("").astype(str)
        if safe.empty:
            return safe

        norm_cols = [_norm(c) for c in safe.columns]
        unnamed_ratio = (
            sum(1 for c in norm_cols if (not c) or c.startswith("unnamed")) / max(len(norm_cols), 1)
        )

        best_row_idx = -1
        best_score = -1
        for row_idx in range(min(len(safe), 12)):
            row_vals = [_norm(v) for v in safe.iloc[row_idx].tolist()]
            non_empty = [v for v in row_vals if v]
            if len(non_empty) < 3:
                continue
            anchor_hits = sum(
                1 for v in non_empty if any(anchor in v for anchor in self.HEADER_ANCHORS)
            )
            score = len(non_empty) + (anchor_hits * 6)
            if score > best_score:
                best_score = score
                best_row_idx = row_idx

        should_promote = best_row_idx >= 0 and (unnamed_ratio >= 0.5 or best_row_idx > 0)
        if should_promote:
            header_raw = [_norm(v) for v in safe.iloc[best_row_idx].tolist()]
            headers: List[str] = []
            seen: Dict[str, int] = {}
            for idx, col_name in enumerate(header_raw):
                base = col_name if (col_name and not col_name.startswith("unnamed")) else f"unnamed_{idx}"
                cnt = seen.get(base, 0)
                headers.append(f"{base}_{cnt}" if cnt else base)
                seen[base] = cnt + 1
            prepared = safe.iloc[best_row_idx + 1 :].reset_index(drop=True).copy()
            prepared.columns = headers
        else:
            prepared = safe.copy()
            prepared.columns = [
                (c if c and not c.startswith("unnamed") else f"unnamed_{idx}")
                for idx, c in enumerate(norm_cols)
            ]

        # Drop fully empty rows after header normalization.
        prepared = prepared[prepared.apply(lambda r: any(str(v).strip() for v in r.tolist()), axis=1)]
        return prepared

    def _extract_join_values(self, values: Dict[str, str]) -> Dict[str, str]:
        out: Dict[str, str] = {}
        for key_name, aliases in self.JOIN_ALIASES.items():
            for alias in aliases:
                value = values.get(_norm(alias), "")
                if value:
                    out[key_name] = value
                    break

        # Generic fallback for any columns that look like file identifiers.
        for col, value in values.items():
            if not value:
                continue
            if "file" in col and "name" in col and "file_name" not in out:
                out["file_name"] = value
            if "path" in col and "file_path" not in out:
                out["file_path"] = value
            if col.endswith("id") and "doc_id" not in out:
                out["doc_id"] = value
        return out

    def _build_request_keys(self, req: TaggingRequest) -> Dict[str, str]:
        metadata = req.metadata if isinstance(req.metadata, dict) else {}
        smart_id = str(metadata.get("smart_id", "") or "").strip()
        file_key = str(metadata.get("file_key", "") or "").strip()
        if not file_key:
            file_key = req.file_hash or (f"file-{req.file_id}" if req.file_id else "")

        file_name = req.file_name or Path(req.file_path or "").name
        return {
            "smart_id": smart_id,
            "file_key": file_key,
            "file_name": file_name,
            "file_path": req.file_path or "",
            "doc_id": req.doc_id or "",
        }

    def _match_score(self, request_keys: Dict[str, str], row_keys: Dict[str, str]) -> int:
        score = 0

        def exact(a: str, b: str) -> bool:
            return bool(a and b and _norm(a) == _norm(b))

        def contains(a: str, b: str) -> bool:
            aa = _norm(a)
            bb = _norm(b)
            return bool(aa and bb and (aa in bb or bb in aa))

        if exact(request_keys.get("smart_id", ""), row_keys.get("smart_id", "")):
            score += 100
        if exact(request_keys.get("file_key", ""), row_keys.get("file_key", "")):
            score += 90
        if exact(request_keys.get("file_name", ""), row_keys.get("file_name", "")):
            score += 80
        elif contains(request_keys.get("file_name", ""), row_keys.get("file_name", "")):
            score += 45

        if exact(request_keys.get("file_path", ""), row_keys.get("file_path", "")):
            score += 70
        elif contains(request_keys.get("file_path", ""), row_keys.get("file_path", "")):
            score += 35

        if exact(request_keys.get("doc_id", ""), row_keys.get("doc_id", "")):
            score += 60

        return score

    @staticmethod
    def _derive_tags(values: Dict[str, str]) -> Dict[str, str]:
        text = " ".join(values.values()).lower()
        derived: Dict[str, str] = {}

        if any(k in text for k in ["invoice", "payment", "amount", "ledger", "account"]):
            derived.setdefault("department", "Finance")
            derived.setdefault("purpose", "Payment")
            derived.setdefault("category", "Invoice")
        if any(k in text for k in ["contract", "agreement", "clause", "legal", "compliance"]):
            derived.setdefault("department", "Legal")
            derived.setdefault("purpose", "Compliance")
            derived.setdefault("category", "Contract")
        if any(k in text for k in ["policy", "procedure", "guideline", "reference"]):
            derived.setdefault("purpose", "Reference")
            derived.setdefault("category", "Policy")
        if "confidential" in text:
            derived["confidentiality"] = "Confidential"
        elif "internal" in text or "restricted" in text:
            derived["confidentiality"] = "Internal"
        elif "public" in text:
            derived["confidentiality"] = "Public"

        return derived


_manager_lock = threading.Lock()
_manager: Optional[MetadataManager] = None


def get_metadata_manager() -> MetadataManager:
    global _manager
    if _manager is None:
        with _manager_lock:
            if _manager is None:
                _manager = MetadataManager()
    return _manager


def set_active_metadata_source(path: str, source: str, *, force: bool = False) -> Tuple[bool, str]:
    return get_metadata_manager().set_active_metadata_source(path=path, source=source, force=force)


def clear_active_metadata_source() -> None:
    get_metadata_manager().clear_active_metadata_source()


def get_metadata_status() -> Dict[str, Any]:
    return get_metadata_manager().get_status()
