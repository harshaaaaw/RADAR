"""
Reporting Manager - Live audit trail and state-matrix export.
"""

from __future__ import annotations

import hashlib
import json
import re
import sqlite3
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from core.config_manager import get_config
from core.logging_manager import get_logger

logger = get_logger("core.reporting_manager")

_LOGIC_OPS = {"AND", "OR"}
_TOKEN_PATTERN = re.compile(r'"[^"]*"|\S+')


def _utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def build_smart_id(file_key: str, department: str = "", when_iso: str = "") -> str:
    """
    Build human-readable deterministic Smart ID.
    Format: [DEPT]-[YYYYMMDD]-[HASH4]
    """
    key = str(file_key or "")
    if not key:
        key = "unknown"

    prefix = re.sub(r"[^A-Z0-9]", "", (department or "DOC").upper())[:3]
    if not prefix:
        prefix = "DOC"

    date_str = ""
    if when_iso:
        date_match = re.search(r"\d{4}-\d{2}-\d{2}", when_iso)
        if date_match:
            date_str = date_match.group(0).replace("-", "")
    if not date_str:
        date_str = datetime.utcnow().strftime("%Y%m%d")

    hash4 = hashlib.sha1(key.encode("utf-8")).hexdigest()[:4].upper()
    return f"{prefix}-{date_str}-{hash4}"


def derive_file_key(file_hash: str = "", file_id: Any = None, file_path: str = "") -> str:
    """
    Stable file key precedence:
    file_hash -> file_id -> file_path.
    """
    if file_hash:
        return str(file_hash)
    if file_id is not None and str(file_id) != "":
        return f"file-{file_id}"
    if file_path:
        return f"path-{file_path}"
    return ""


def normalize_file_type(file_type: str = "", file_name: str = "", file_path: str = "") -> str:
    candidate = (file_type or "").strip().lower().lstrip(".")
    if candidate:
        return candidate

    path = file_name or file_path
    if path:
        suffix = Path(path).suffix.lower().lstrip(".")
        return suffix
    return ""


@dataclass
class AuditEvent:
    event_time: str = ""
    file_key: str = ""
    smart_id: str = ""
    file_name: str = ""
    file_path: str = ""
    stage: str = ""
    status: str = ""
    worker_id: str = ""
    file_type: str = ""
    error_type: str = ""
    error_message: str = ""
    payload_json: Any = None


@dataclass
class FileStateRow:
    file_key: str
    smart_id: str = ""
    file_name: str = ""
    category: str = ""
    department: str = ""
    purpose: str = ""
    key_names: str = ""
    amount_found: str = ""
    important_dates: str = ""
    location_mentioned: str = ""
    confidentiality: str = ""
    current_status: str = ""
    processed_on: str = ""
    file_type: str = ""
    file_size: int = 0
    file_path: str = ""
    updated_at: str = ""
    tag_confidence: float = 0.0
    source_stage: str = ""
    worker_id: str = ""
    tagging_status: str = ""
    review_required: bool = False
    tagger_version: str = ""
    taxonomy_version: str = ""
    confidence_json: str = ""
    extended_metadata_json: str = ""
    constraint_source: str = ""
    forced_flag: bool = False
    original_label: str = ""
    original_score: float = 0.0
    match_mode: str = ""
    constraint_version: str = ""
    # ===== 12 Core Taxonomy Dimensions =====
    metadata_level_code: str = "File"
    record_class_name: str = "Undefined"
    record_category_name_functional: str = ""
    record_category_name_transactional: str = ""
    record_type_code: str = ""
    business_unit_name: str = ""
    sub_business_unit_name: str = ""
    iso_country_code: str = ""
    record_format_name: str = "Electronic"
    original_record_location_type_name: str = "Shared Drive"
    data_classification_name: str = "GE Internal"
    divestiture_deal_name: str = ""
    dynamic_subtags: str = ""
    # ===== Accuracy Metrics =====
    pipeline_type: str = ""
    extraction_accuracy: float = 0.0
    text_area_pct: float = 0.0
    non_text_area_pct: float = 0.0
    raw_char_count: int = 0
    processed_char_count: int = 0
    preprocessing_gain_pct: float = 0.0
    accuracy_loss_json: str = ""
    page_metrics_json: str = ""
    accuracy_tier: str = ""



class _ReportingManager:
    EXPORT_DEFAULTS = {
        "smart_id": "DOC-00000000-0000",
        "file_name": "unknown_file",
        "category": "General",
        "department": "Operations",
        "purpose": "Reference",
        "key_names": "None",
        "amount_found": "None",
        "important_dates": "None",
        "location_mentioned": "None",
        "confidentiality": "Public",
        "current_status": "unknown",
        "processed_on": "1970-01-01T00:00:00Z",
        "file_type": "unknown",
        "file_size": 0,
        # 12-Dimension defaults
        "metadata_level_code": "File",
        "record_class_name": "Undefined",
        "record_category_name_functional": "",
        "record_category_name_transactional": "",
        "record_type_code": "",
        "business_unit_name": "",
        "sub_business_unit_name": "",
        "iso_country_code": "",
        "record_format_name": "Electronic",
        "original_record_location_type_name": "Shared Drive",
        "data_classification_name": "GE Internal",
        "divestiture_deal_name": "",
        "dynamic_subtags": "",
        # Accuracy metrics defaults
        "pipeline_type": "",
        "extraction_accuracy": 0.0,
        "text_area_pct": 0.0,
        "non_text_area_pct": 0.0,
        "accuracy_tier": "",
    }


    def __init__(self) -> None:
        cfg = get_config()
        self.config = cfg
        self._config = cfg
        self.audit_dir = Path(cfg.paths.working_root) / "audit"
        self.audit_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = self.audit_dir / "audit.db"
        self._schema_lock = threading.Lock()
        self._schema_ready = False

        self._connect_timeout = 20.0
        self._busy_timeout_ms = 10000
        self._max_retries = 6
        self._retry_base_seconds = 0.05

        self._ensure_schema()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(
            str(self.db_path),
            timeout=self._connect_timeout,
            check_same_thread=False,
        )
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute(f"PRAGMA busy_timeout={self._busy_timeout_ms}")
        return conn

    def _execute_with_retry(self, fn, *, commit: bool) -> Any:
        last_error: Optional[Exception] = None
        for attempt in range(self._max_retries):
            conn: Optional[sqlite3.Connection] = None
            try:
                conn = self._connect()
                result = fn(conn)
                if commit:
                    conn.commit()
                return result
            except sqlite3.OperationalError as exc:
                last_error = exc
                if "locked" in str(exc).lower() or "busy" in str(exc).lower():
                    sleep_s = self._retry_base_seconds * (2**attempt)
                    time.sleep(min(sleep_s, 1.0))
                    continue
                raise
            finally:
                if conn is not None:
                    conn.close()
        if last_error:
            raise last_error
        raise RuntimeError("Unknown sqlite execution failure")

    def _ensure_schema(self) -> None:
        if self._schema_ready:
            return
        with self._schema_lock:
            if self._schema_ready:
                return

            def _create_schema(conn: sqlite3.Connection) -> None:
                cur = conn.cursor()
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS audit_events (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        event_time TEXT NOT NULL,
                        file_key TEXT,
                        smart_id TEXT,
                        file_name TEXT,
                        file_path TEXT,
                        stage TEXT,
                        status TEXT,
                        worker_id TEXT,
                        file_type TEXT,
                        error_type TEXT,
                        error_message TEXT,
                        payload_json TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_events_time ON audit_events (event_time DESC)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_events_status ON audit_events (status)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_events_stage ON audit_events (stage)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_events_type ON audit_events (file_type)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_events_file_key ON audit_events (file_key)")

                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS file_state (
                        file_key TEXT PRIMARY KEY,
                        smart_id TEXT,
                        file_name TEXT,
                        category TEXT,
                        department TEXT,
                        purpose TEXT,
                        key_names TEXT,
                        amount_found TEXT,
                        important_dates TEXT,
                        location_mentioned TEXT,
                        confidentiality TEXT,
                        current_status TEXT,
                        processed_on TEXT,
                        file_type TEXT,
                        file_size INTEGER,
                        file_path TEXT,
                        updated_at TEXT,
                        tag_confidence REAL,
                        source_stage TEXT,
                        worker_id TEXT,
                        tagging_status TEXT,
                        review_required INTEGER DEFAULT 0,
                        tagger_version TEXT,
                        taxonomy_version TEXT,
                        confidence_json TEXT,
                        extended_metadata_json TEXT,
                        constraint_source TEXT,
                        forced_flag INTEGER DEFAULT 0,
                        original_label TEXT,
                        original_score REAL DEFAULT 0.0,
                        match_mode TEXT,
                        constraint_version TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_status ON file_state (current_status)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_stage ON file_state (source_stage)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_type ON file_state (file_type)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_smart_id ON file_state (smart_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_worker ON file_state (worker_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_updated_at ON file_state (updated_at DESC)")

                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS tag_feedback (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        file_key TEXT,
                        smart_id TEXT,
                        field_name TEXT,
                        old_value TEXT,
                        new_value TEXT,
                        actor TEXT,
                        event_time TEXT,
                        reason TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_tag_feedback_field ON tag_feedback (field_name)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_tag_feedback_new_value ON tag_feedback (new_value)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_tag_feedback_time ON tag_feedback (event_time DESC)")

                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS taxonomy_versions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        version_id TEXT UNIQUE,
                        source_file TEXT,
                        checksum TEXT,
                        loaded_at TEXT,
                        status TEXT,
                        notes TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_taxonomy_versions_loaded_at ON taxonomy_versions (loaded_at DESC)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_taxonomy_versions_status ON taxonomy_versions (status)")

                # Backward-compatible migration for existing DBs.
                columns = {row["name"] for row in cur.execute("PRAGMA table_info(file_state)").fetchall()}
                if "worker_id" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN worker_id TEXT")
                if "tagging_status" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN tagging_status TEXT")
                if "review_required" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN review_required INTEGER DEFAULT 0")
                if "tagger_version" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN tagger_version TEXT")
                if "taxonomy_version" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN taxonomy_version TEXT")
                if "confidence_json" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN confidence_json TEXT")
                if "extended_metadata_json" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN extended_metadata_json TEXT")
                if "constraint_source" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN constraint_source TEXT DEFAULT ''")
                if "forced_flag" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN forced_flag INTEGER DEFAULT 0")
                if "original_label" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN original_label TEXT DEFAULT ''")
                if "original_score" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN original_score REAL DEFAULT 0.0")
                if "match_mode" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN match_mode TEXT DEFAULT ''")
                if "constraint_version" not in columns:
                    cur.execute("ALTER TABLE file_state ADD COLUMN constraint_version TEXT DEFAULT ''")
                # ===== 12-Dimension Schema Migration =====
                dim_migrations = {
                    "metadata_level_code": "TEXT DEFAULT 'File'",
                    "record_class_name": "TEXT DEFAULT 'Undefined'",
                    "record_category_name_functional": "TEXT DEFAULT ''",
                    "record_category_name_transactional": "TEXT DEFAULT ''",
                    "record_type_code": "TEXT DEFAULT ''",
                    "business_unit_name": "TEXT DEFAULT ''",
                    "sub_business_unit_name": "TEXT DEFAULT ''",
                    "iso_country_code": "TEXT DEFAULT ''",
                    "record_format_name": "TEXT DEFAULT 'Electronic'",
                    "original_record_location_type_name": "TEXT DEFAULT 'Shared Drive'",
                    "data_classification_name": "TEXT DEFAULT 'GE Internal'",
                    "divestiture_deal_name": "TEXT DEFAULT ''",
                    "dynamic_subtags": "TEXT DEFAULT ''",
                }
                for col_name, col_type in dim_migrations.items():
                    if col_name not in columns:
                        cur.execute(f"ALTER TABLE file_state ADD COLUMN {col_name} {col_type}")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_tagging_status ON file_state (tagging_status)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_file_state_review_required ON file_state (review_required)")
                # 12-Dimension performance indices
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_bu ON file_state (business_unit_name)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_functional_cat ON file_state (record_category_name_functional)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_deal ON file_state (divestiture_deal_name)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_country ON file_state (iso_country_code)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_record_class ON file_state (record_class_name)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_fs_data_class ON file_state (data_classification_name)")

                # ===== Accuracy Metrics Schema Migration =====
                accuracy_migrations = {
                    "pipeline_type": "TEXT DEFAULT ''",
                    "extraction_accuracy": "REAL DEFAULT NULL",
                    "enhanced_accuracy": "REAL DEFAULT NULL",
                    "approval_status": "TEXT DEFAULT 'Full Baseline'",
                    "text_area_pct": "REAL DEFAULT NULL",
                    "non_text_area_pct": "REAL DEFAULT NULL",
                    "raw_char_count": "INTEGER DEFAULT NULL",
                    "processed_char_count": "INTEGER DEFAULT NULL",
                    "preprocessing_gain_pct": "REAL DEFAULT NULL",
                    "accuracy_loss_json": "TEXT DEFAULT ''",
                    "page_metrics_json": "TEXT DEFAULT ''",
                    "accuracy_tier": "TEXT DEFAULT ''",
                }
                for col_name, col_type in accuracy_migrations.items():
                    if col_name not in columns:
                        cur.execute(f"ALTER TABLE file_state ADD COLUMN {col_name} {col_type}")

                # Create snippet_reviews table for HITL visual audits
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS snippet_reviews (
                        review_id TEXT PRIMARY KEY,
                        smart_id TEXT NOT NULL,
                        page_num INTEGER NOT NULL,
                        snippet_type TEXT NOT NULL,
                        snippet_path TEXT NOT NULL,
                        bounding_box_json TEXT NOT NULL,
                        accuracy_impact REAL NOT NULL,
                        reviewer_role TEXT NOT NULL,
                        status TEXT DEFAULT 'pending',
                        feature_vector_path TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_snippet_reviews_smart_id ON snippet_reviews (smart_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_snippet_reviews_status ON snippet_reviews (status)")

                # Backward-compatible migrations for snippet_reviews
                sr_columns = {row["name"] for row in cur.execute("PRAGMA table_info(snippet_reviews)").fetchall()}
                if "reviewed_at" not in sr_columns:
                    cur.execute("ALTER TABLE snippet_reviews ADD COLUMN reviewed_at TEXT")
                if "reviewed_by" not in sr_columns:
                    cur.execute("ALTER TABLE snippet_reviews ADD COLUMN reviewed_by TEXT")
                if "review_reason" not in sr_columns:
                    cur.execute("ALTER TABLE snippet_reviews ADD COLUMN review_reason TEXT")
                if "file_size_bytes" not in sr_columns:
                    cur.execute("ALTER TABLE snippet_reviews ADD COLUMN file_size_bytes INTEGER DEFAULT 0")

                # Create review activity log table for audit trails
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS review_activity_log (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        review_id TEXT NOT NULL,
                        smart_id TEXT NOT NULL,
                        action TEXT NOT NULL,
                        actor TEXT DEFAULT 'Dashboard User',
                        reason TEXT DEFAULT '',
                        timestamp TEXT NOT NULL,
                        accuracy_before REAL,
                        accuracy_after REAL,
                        snippet_type TEXT,
                        file_name TEXT
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_review_log_smart_id ON review_activity_log (smart_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_review_log_timestamp ON review_activity_log (timestamp DESC)")

            self._execute_with_retry(_create_schema, commit=True)
            self._schema_ready = True


    @staticmethod
    def _coerce_payload_json(payload: Any) -> str:
        if payload is None:
            return "{}"
        if isinstance(payload, str):
            return payload
        try:
            return json.dumps(payload, ensure_ascii=False, default=str)
        except Exception:
            return "{}"

    @staticmethod
    def _join_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, list):
            return ", ".join(str(v) for v in value if v is not None)
        return str(value)

    def record_event(self, event: AuditEvent) -> None:
        self._ensure_schema()
        event_time = event.event_time or _utc_now_iso()
        file_type = normalize_file_type(event.file_type, event.file_name, event.file_path)
        payload = self._coerce_payload_json(event.payload_json)

        def _insert(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                INSERT INTO audit_events (
                    event_time, file_key, smart_id, file_name, file_path,
                    stage, status, worker_id, file_type, error_type,
                    error_message, payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_time,
                    event.file_key,
                    event.smart_id,
                    event.file_name,
                    event.file_path,
                    event.stage,
                    event.status,
                    event.worker_id,
                    file_type,
                    event.error_type,
                    event.error_message,
                    payload,
                ),
            )

        self._execute_with_retry(_insert, commit=True)

    def upsert_file_state(self, state: FileStateRow) -> None:
        self._ensure_schema()
        if not state.file_key:
            raise ValueError("file_state upsert requires file_key")

        state = self._sanitize_state_row(state)

        processed_on = state.processed_on or _utc_now_iso()
        updated_at = state.updated_at or _utc_now_iso()
        file_type = normalize_file_type(state.file_type, state.file_name, state.file_path)
        smart_id = state.smart_id or build_smart_id(state.file_key, state.department, processed_on)

        def _upsert(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                INSERT INTO file_state (
                    file_key, smart_id, file_name, category, department, purpose, key_names,
                    amount_found, important_dates, location_mentioned, confidentiality,
                    current_status, processed_on, file_type, file_size, file_path,
                    updated_at, tag_confidence, source_stage, worker_id,
                    tagging_status, review_required, tagger_version, taxonomy_version, confidence_json,
                    extended_metadata_json, constraint_source, forced_flag, original_label, original_score,
                    match_mode, constraint_version,
                    metadata_level_code, record_class_name, record_category_name_functional,
                    record_category_name_transactional, record_type_code, business_unit_name,
                    sub_business_unit_name, iso_country_code, record_format_name,
                    original_record_location_type_name, data_classification_name, divestiture_deal_name,
                    dynamic_subtags
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(file_key) DO UPDATE SET
                    smart_id=CASE
                        WHEN excluded.smart_id IS NOT NULL AND TRIM(excluded.smart_id) != '' THEN excluded.smart_id
                        ELSE file_state.smart_id
                    END,
                    file_name=CASE
                        WHEN excluded.file_name IS NOT NULL AND TRIM(excluded.file_name) != '' THEN excluded.file_name
                        ELSE file_state.file_name
                    END,
                    category=CASE
                        WHEN excluded.category IS NOT NULL AND TRIM(excluded.category) != '' THEN excluded.category
                        ELSE file_state.category
                    END,
                    department=CASE
                        WHEN excluded.department IS NOT NULL AND TRIM(excluded.department) != '' THEN excluded.department
                        ELSE file_state.department
                    END,
                    purpose=CASE
                        WHEN excluded.purpose IS NOT NULL AND TRIM(excluded.purpose) != '' THEN excluded.purpose
                        ELSE file_state.purpose
                    END,
                    key_names=CASE
                        WHEN excluded.key_names IS NOT NULL AND TRIM(excluded.key_names) != '' THEN excluded.key_names
                        ELSE file_state.key_names
                    END,
                    amount_found=CASE
                        WHEN excluded.amount_found IS NOT NULL AND TRIM(excluded.amount_found) != '' THEN excluded.amount_found
                        ELSE file_state.amount_found
                    END,
                    important_dates=CASE
                        WHEN excluded.important_dates IS NOT NULL AND TRIM(excluded.important_dates) != '' THEN excluded.important_dates
                        ELSE file_state.important_dates
                    END,
                    location_mentioned=CASE
                        WHEN excluded.location_mentioned IS NOT NULL AND TRIM(excluded.location_mentioned) != '' THEN excluded.location_mentioned
                        ELSE file_state.location_mentioned
                    END,
                    confidentiality=CASE
                        WHEN excluded.confidentiality IS NOT NULL AND TRIM(excluded.confidentiality) != '' THEN excluded.confidentiality
                        ELSE file_state.confidentiality
                    END,
                    current_status=excluded.current_status,
                    processed_on=excluded.processed_on,
                    file_type=CASE
                        WHEN excluded.file_type IS NOT NULL AND TRIM(excluded.file_type) != '' THEN excluded.file_type
                        ELSE file_state.file_type
                    END,
                    file_size=CASE
                        WHEN excluded.file_size IS NOT NULL AND excluded.file_size > 0 THEN excluded.file_size
                        ELSE file_state.file_size
                    END,
                    file_path=CASE
                        WHEN excluded.file_path IS NOT NULL AND TRIM(excluded.file_path) != '' THEN excluded.file_path
                        ELSE file_state.file_path
                    END,
                    updated_at=excluded.updated_at,
                    tag_confidence=CASE
                        WHEN excluded.tag_confidence IS NOT NULL AND excluded.tag_confidence > 0 THEN excluded.tag_confidence
                        ELSE file_state.tag_confidence
                    END,
                    source_stage=excluded.source_stage,
                    worker_id=CASE
                        WHEN excluded.worker_id IS NOT NULL AND TRIM(excluded.worker_id) != '' THEN excluded.worker_id
                        ELSE file_state.worker_id
                    END,
                    tagging_status=CASE
                        WHEN excluded.tagging_status IS NOT NULL AND TRIM(excluded.tagging_status) != '' THEN excluded.tagging_status
                        ELSE file_state.tagging_status
                    END,
                    review_required=CASE
                        WHEN excluded.review_required IS NOT NULL THEN excluded.review_required
                        ELSE file_state.review_required
                    END,
                    tagger_version=CASE
                        WHEN excluded.tagger_version IS NOT NULL AND TRIM(excluded.tagger_version) != '' THEN excluded.tagger_version
                        ELSE file_state.tagger_version
                    END,
                    taxonomy_version=CASE
                        WHEN excluded.taxonomy_version IS NOT NULL AND TRIM(excluded.taxonomy_version) != '' THEN excluded.taxonomy_version
                        ELSE file_state.taxonomy_version
                    END,
                    confidence_json=CASE
                        WHEN excluded.confidence_json IS NOT NULL AND TRIM(excluded.confidence_json) != '' THEN excluded.confidence_json
                        ELSE file_state.confidence_json
                    END,
                    extended_metadata_json=CASE
                        WHEN excluded.extended_metadata_json IS NOT NULL AND TRIM(excluded.extended_metadata_json) != '' THEN excluded.extended_metadata_json
                        ELSE file_state.extended_metadata_json
                    END,
                    constraint_source=CASE
                        WHEN excluded.constraint_source IS NOT NULL AND TRIM(excluded.constraint_source) != '' THEN excluded.constraint_source
                        ELSE file_state.constraint_source
                    END,
                    forced_flag=CASE
                        WHEN excluded.forced_flag IS NOT NULL THEN excluded.forced_flag
                        ELSE file_state.forced_flag
                    END,
                    original_label=CASE
                        WHEN excluded.original_label IS NOT NULL AND TRIM(excluded.original_label) != '' THEN excluded.original_label
                        ELSE file_state.original_label
                    END,
                    original_score=CASE
                        WHEN excluded.original_score IS NOT NULL AND excluded.original_score > 0 THEN excluded.original_score
                        ELSE file_state.original_score
                    END,
                    match_mode=CASE
                        WHEN excluded.match_mode IS NOT NULL AND TRIM(excluded.match_mode) != '' THEN excluded.match_mode
                        ELSE file_state.match_mode
                    END,
                    constraint_version=CASE
                        WHEN excluded.constraint_version IS NOT NULL AND TRIM(excluded.constraint_version) != '' THEN excluded.constraint_version
                        ELSE file_state.constraint_version
                    END,
                    metadata_level_code=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.metadata_level_code ELSE COALESCE(NULLIF(TRIM(excluded.metadata_level_code), ''), file_state.metadata_level_code) END,
                    record_class_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.record_class_name ELSE COALESCE(NULLIF(TRIM(excluded.record_class_name), ''), file_state.record_class_name) END,
                    record_category_name_functional=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.record_category_name_functional ELSE COALESCE(NULLIF(TRIM(excluded.record_category_name_functional), ''), file_state.record_category_name_functional) END,
                    record_category_name_transactional=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.record_category_name_transactional ELSE COALESCE(NULLIF(TRIM(excluded.record_category_name_transactional), ''), file_state.record_category_name_transactional) END,
                    record_type_code=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.record_type_code ELSE COALESCE(NULLIF(TRIM(excluded.record_type_code), ''), file_state.record_type_code) END,
                    business_unit_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.business_unit_name ELSE COALESCE(NULLIF(TRIM(excluded.business_unit_name), ''), file_state.business_unit_name) END,
                    sub_business_unit_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.sub_business_unit_name ELSE COALESCE(NULLIF(TRIM(excluded.sub_business_unit_name), ''), file_state.sub_business_unit_name) END,
                    iso_country_code=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.iso_country_code ELSE COALESCE(NULLIF(TRIM(excluded.iso_country_code), ''), file_state.iso_country_code) END,
                    record_format_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.record_format_name ELSE COALESCE(NULLIF(TRIM(excluded.record_format_name), ''), file_state.record_format_name) END,
                    original_record_location_type_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.original_record_location_type_name ELSE COALESCE(NULLIF(TRIM(excluded.original_record_location_type_name), ''), file_state.original_record_location_type_name) END,
                    data_classification_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.data_classification_name ELSE COALESCE(NULLIF(TRIM(excluded.data_classification_name), ''), file_state.data_classification_name) END,
                    divestiture_deal_name=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.divestiture_deal_name ELSE COALESCE(NULLIF(TRIM(excluded.divestiture_deal_name), ''), file_state.divestiture_deal_name) END,
                    dynamic_subtags=CASE WHEN excluded.source_stage = 'tagging' THEN excluded.dynamic_subtags ELSE COALESCE(NULLIF(TRIM(excluded.dynamic_subtags), ''), file_state.dynamic_subtags) END
                """,
                (
                    state.file_key,
                    smart_id,
                    state.file_name,
                    state.category,
                    state.department,
                    state.purpose,
                    self._join_value(state.key_names),
                    self._join_value(state.amount_found),
                    self._join_value(state.important_dates),
                    self._join_value(state.location_mentioned),
                    self._join_value(state.confidentiality),
                    state.current_status,
                    processed_on,
                    file_type,
                    int(state.file_size or 0),
                    state.file_path,
                    updated_at,
                    float(state.tag_confidence or 0.0),
                    state.source_stage,
                    state.worker_id,
                    str(state.tagging_status or ""),
                    1 if bool(state.review_required) else 0,
                    str(state.tagger_version or ""),
                    str(state.taxonomy_version or ""),
                    str(state.confidence_json or ""),
                    str(state.extended_metadata_json or ""),
                    str(state.constraint_source or ""),
                    1 if bool(state.forced_flag) else 0,
                    str(state.original_label or ""),
                    float(state.original_score or 0.0),
                    str(state.match_mode or ""),
                    str(state.constraint_version or ""),
                    # 12 Taxonomy Dimensions
                    str(state.metadata_level_code or "File"),
                    str(state.record_class_name or "Undefined"),
                    str(state.record_category_name_functional or ""),
                    str(state.record_category_name_transactional or ""),
                    str(state.record_type_code or ""),
                    str(state.business_unit_name or ""),
                    str(state.sub_business_unit_name or ""),
                    str(state.iso_country_code or ""),
                    str(state.record_format_name or "Electronic"),
                    str(state.original_record_location_type_name or "Shared Drive"),
                    str(state.data_classification_name or "GE Internal"),
                    str(state.divestiture_deal_name or ""),
                    self._join_value(state.dynamic_subtags),
                ),
            )

        self._execute_with_retry(_upsert, commit=True)

    def update_accuracy_metrics(self, file_key: str, metrics: Dict[str, Any]) -> None:
        """Update accuracy metrics for a document (separate from main upsert)."""
        self._ensure_schema()
        if not file_key:
            return

        def _update(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                UPDATE file_state SET
                    pipeline_type = ?,
                    extraction_accuracy = ?,
                    enhanced_accuracy = COALESCE(enhanced_accuracy, ?),
                    approval_status = COALESCE(approval_status, 'Full Baseline'),
                    text_area_pct = ?,
                    non_text_area_pct = ?,
                    raw_char_count = ?,
                    processed_char_count = ?,
                    preprocessing_gain_pct = ?,
                    accuracy_loss_json = ?,
                    page_metrics_json = ?,
                    accuracy_tier = ?
                WHERE file_key = ?
                """,
                (
                    str(metrics.get("pipeline_type", "") or ""),
                    float(metrics.get("extraction_accuracy") or 0.0),
                    float(metrics.get("extraction_accuracy") or 0.0),  # Default enhanced accuracy to extraction accuracy
                    float(metrics.get("text_area_pct") or 0.0),
                    float(metrics.get("non_text_area_pct") or 0.0),
                    int(metrics.get("raw_char_count") or 0),
                    int(metrics.get("processed_char_count") or 0),
                    float(metrics.get("preprocessing_gain_pct") or 0.0),
                    str(metrics.get("accuracy_loss_json", "") or ""),
                    str(metrics.get("page_metrics_json", "") or ""),
                    str(metrics.get("accuracy_tier", "") or ""),
                    file_key,
                ),
            )

        self._execute_with_retry(_update, commit=True)

    def _sanitize_state_row(self, state: FileStateRow) -> FileStateRow:
        """Guarantee deterministic non-empty required fields before persistence."""
        # Core tagging fields
        if not str(state.category or "").strip():
            state.category = str(self.EXPORT_DEFAULTS["category"])
        if not str(state.department or "").strip():
            state.department = str(self.EXPORT_DEFAULTS["department"])
        if not str(state.purpose or "").strip():
            state.purpose = str(self.EXPORT_DEFAULTS["purpose"])
        if not str(state.confidentiality or "").strip():
            state.confidentiality = str(self.EXPORT_DEFAULTS["confidentiality"])

        # Export-visible non-empty fields
        if not str(state.file_name or "").strip():
            state.file_name = str(self.EXPORT_DEFAULTS["file_name"])
        if not str(state.smart_id or "").strip():
            state.smart_id = str(self.EXPORT_DEFAULTS["smart_id"])
        if not str(state.current_status or "").strip():
            state.current_status = str(self.EXPORT_DEFAULTS["current_status"])
        if not str(state.processed_on or "").strip():
            state.processed_on = str(self.EXPORT_DEFAULTS["processed_on"])
        if not str(state.file_type or "").strip():
            state.file_type = str(self.EXPORT_DEFAULTS["file_type"])

        # Coerce list-like fields to non-empty placeholders.
        if not self._join_value(state.key_names).strip():
            state.key_names = str(self.EXPORT_DEFAULTS["key_names"])
        if not self._join_value(state.amount_found).strip():
            state.amount_found = str(self.EXPORT_DEFAULTS["amount_found"])
        if not self._join_value(state.important_dates).strip():
            state.important_dates = str(self.EXPORT_DEFAULTS["important_dates"])
        if not self._join_value(state.location_mentioned).strip():
            state.location_mentioned = str(self.EXPORT_DEFAULTS["location_mentioned"])

        return state

    def _validate_export_row(
        self,
        row: Dict[str, Any],
        sheet3_registry: Optional[Dict[str, Any]] = None,
    ) -> List[str]:
        """
        Pre-export strict validator (Phase 4, plan item).
        Validates a single export row:
        1. All required fields must be non-empty (uses EXPORT_DEFAULTS for fallback, logs WARNING).
        2. category, department, confidentiality must be in Sheet 3 allowed set (logs WARNING).
        Does NOT silently mask — always writes with a WARNING log for operator visibility.
        Returns list of warnings (empty if row is compliant).
        """
        warnings: List[str] = []

        # 1. Check required fields non-empty
        for field_key in ("smart_id", "file_name", "category", "department", "purpose",
                          "confidentiality", "current_status", "processed_on", "file_type"):
            val = str(row.get(field_key, "") or "").strip()
            if not val or val.lower() in ("nan", "none", ""):
                default = self.EXPORT_DEFAULTS.get(field_key, "")
                row[field_key] = default
                msg = f"Export row missing '{field_key}', defaulted to '{default}'"
                warnings.append(msg)
                logger.warning(msg)

        # 2. Validate constrained fields against Sheet 3
        if sheet3_registry:
            constraint_map = {
                "category": "record_category_name_functional",
                "department": "business_unit_name",
                "confidentiality": "data_classification_name",
            }
            for field, registry_key in constraint_map.items():
                allowed_set = sheet3_registry.get(registry_key)
                if not allowed_set:
                    continue
                val = str(row.get(field, "") or "").strip()
                if val and val.lower() not in {a.strip().lower() for a in allowed_set}:
                    msg = f"Export row '{field}'='{val}' not in Sheet 3 allowed set for '{registry_key}'"
                    warnings.append(msg)
                    logger.warning(msg)

        return warnings


    @staticmethod
    def _tokenize_filter_query(filter_query: str) -> List[str]:
        return _TOKEN_PATTERN.findall(filter_query or "")

    def _parse_term(
        self,
        token: str,
        field_map: Dict[str, str],
        *,
        text_fields: Optional[List[str]] = None,
    ) -> Tuple[str, List[Any]]:
        text_fields = text_fields or []
        token = token.strip()
        if not token:
            raise ValueError("Empty token in filter expression")

        if ":" in token:
            field, raw_val = token.split(":", 1)
            field = field.strip().lower()
            value = raw_val.strip().strip('"')
            if not value:
                raise ValueError(f"Missing value for filter field: {field}")
            if field not in field_map:
                raise ValueError(f"Unsupported filter field: {field}")

            col = field_map[field]
            if field in {"name", "path"}:
                return f"LOWER({col}) LIKE LOWER(?)", [f"%{value}%"]
            if field == "type":
                value_norm = value.lstrip(".")
                return f"(LOWER({col}) = LOWER(?) OR LOWER({col}) = LOWER(?))", [value_norm, f".{value_norm}"]
            return f"LOWER({col}) = LOWER(?)", [value]

        # Free-text fallback
        if not text_fields:
            raise ValueError(f"Invalid token (expected field:value): {token}")
        val = token.strip('"')
        likes = [f"LOWER({col}) LIKE LOWER(?)" for col in text_fields]
        return "(" + " OR ".join(likes) + ")", [f"%{val}%"] * len(text_fields)

    def _build_filter_sql(
        self,
        filter_query: str,
        field_map: Dict[str, str],
        *,
        text_fields: Optional[List[str]] = None,
    ) -> Tuple[str, List[Any]]:
        tokens = self._tokenize_filter_query(filter_query)
        if not tokens:
            return "", []

        clauses: List[str] = []
        params: List[Any] = []
        connectors: List[str] = []
        expect_term = True

        for token in tokens:
            op = token.upper()
            if op in _LOGIC_OPS:
                if expect_term:
                    raise ValueError("Invalid filter expression: operator without term")
                connectors.append(op)
                expect_term = True
                continue

            if not expect_term:
                connectors.append("AND")

            clause, values = self._parse_term(token, field_map, text_fields=text_fields)
            clauses.append(clause)
            params.extend(values)
            expect_term = False

        if expect_term and clauses:
            raise ValueError("Invalid filter expression: trailing operator")

        if not clauses:
            return "", []

        sql = clauses[0]
        for idx in range(1, len(clauses)):
            sql = f"({sql}) {connectors[idx - 1]} ({clauses[idx]})"
        return sql, params

    def search_events(self, filter_query: str, limit: int = 50) -> List[Dict[str, Any]]:
        self._ensure_schema()
        safe_limit = max(1, min(int(limit or 50), 1000))

        if not (filter_query or "").strip():
            return self.get_live_feed(safe_limit)

        field_map = {
            "status": "status",
            "stage": "stage",
            "type": "file_type",
            "worker": "worker_id",
            "name": "file_name",
            "path": "file_path",
            "smart_id": "smart_id",
        }
        where_sql, params = self._build_filter_sql(
            filter_query,
            field_map,
            text_fields=["file_name", "file_path", "error_message", "payload_json"],
        )

        query = """
            SELECT
                id, event_time, file_key, smart_id, file_name, file_path, stage,
                status, worker_id, file_type, error_type, error_message, payload_json
            FROM audit_events
        """
        if where_sql:
            query += f" WHERE {where_sql}"
        query += " ORDER BY id DESC LIMIT ?"
        params.append(safe_limit)

        def _search(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(query, tuple(params)).fetchall()
            return [dict(row) for row in rows]

        return self._execute_with_retry(_search, commit=False)

    def get_live_feed(self, limit: int = 50) -> List[Dict[str, Any]]:
        self._ensure_schema()
        safe_limit = max(1, min(int(limit or 50), 1000))

        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT
                    id, event_time, file_key, smart_id, file_name, file_path, stage,
                    status, worker_id, file_type, error_type, error_message, payload_json
                FROM audit_events
                ORDER BY id DESC
                LIMIT ?
                """,
                (safe_limit,),
            ).fetchall()
            return [dict(row) for row in rows]

        return self._execute_with_retry(_read, commit=False)

    def export_state_matrix_xlsx(self, filters: Optional[Dict[str, Any]], out_path: str) -> str:
        self._ensure_schema()
        filters = filters or {}
        clauses: List[str] = []
        params: List[Any] = []

        filter_query = str(filters.get("filter_query") or "").strip()
        if filter_query:
            field_map = {
                "status": "current_status",
                "stage": "source_stage",
                "type": "file_type",
                "worker": "worker_id",
                "name": "file_name",
                "path": "file_path",
                "smart_id": "smart_id",
                # 12-Dimension filter fields
                "bu": "business_unit_name",
                "business_unit": "business_unit_name",
                "country": "iso_country_code",
                "deal": "divestiture_deal_name",
                "record_class": "record_class_name",
                "functional": "record_category_name_functional",
                "transactional": "record_category_name_transactional",
                "data_class": "data_classification_name",
            }
            where_sql, where_params = self._build_filter_sql(
                filter_query,
                field_map,
                text_fields=["file_name", "file_path", "purpose", "department", "category",
                             "business_unit_name", "divestiture_deal_name"],
            )
            if where_sql:
                clauses.append(where_sql)
                params.extend(where_params)

        simple_map = {
            "status": "current_status",
            "stage": "source_stage",
            "type": "file_type",
            "worker": "worker_id",
            "name": "file_name",
            "path": "file_path",
            "smart_id": "smart_id",
            "department": "department",
            "category": "category",
            "purpose": "purpose",
            # 12-Dimension direct filters
            "business_unit": "business_unit_name",
            "country": "iso_country_code",
            "deal": "divestiture_deal_name",
            "record_class": "record_class_name",
            "data_classification": "data_classification_name",
        }
        for key, column in simple_map.items():
            if key == "filter_query":
                continue
            raw_val = filters.get(key)
            if raw_val is None or str(raw_val).strip() == "":
                continue
            value = str(raw_val).strip()
            if key in {"name", "path"}:
                clauses.append(f"LOWER({column}) LIKE LOWER(?)")
                params.append(f"%{value}%")
            elif key == "type":
                value = value.lstrip(".")
                clauses.append(f"(LOWER({column}) = LOWER(?) OR LOWER({column}) = LOWER(?))")
                params.extend([value, f".{value}"])
            else:
                clauses.append(f"LOWER({column}) = LOWER(?)")
                params.append(value)
        sql = """
            SELECT
                smart_id,
                file_name,
                file_path,
                metadata_level_code,
                record_class_name,
                record_category_name_functional,
                record_category_name_transactional,
                record_type_code,
                business_unit_name,
                sub_business_unit_name,
                iso_country_code,
                record_format_name,
                original_record_location_type_name,
                data_classification_name,
                divestiture_deal_name,
                key_names,
                amount_found,
                important_dates,
                location_mentioned,
                dynamic_subtags,
                file_type,
                file_size,
                updated_at,
                tag_confidence,
                pipeline_type,
                extraction_accuracy,
                enhanced_accuracy,
                approval_status,
                text_area_pct,
                non_text_area_pct,
                preprocessing_gain_pct,
                accuracy_loss_json
            FROM file_state
        """
        if clauses:
            sql += " WHERE " + " AND ".join(f"({c})" for c in clauses)
        sql += " ORDER BY updated_at DESC"

        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(sql, tuple(params)).fetchall()
            return [dict(row) for row in rows]

        rows = self._execute_with_retry(_read, commit=False)

        export_path = out_path.strip() if out_path else ""
        if not export_path:
            export_path = str(self.audit_dir / f"state_matrix_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
        elif Path(export_path).is_dir():
            export_path = str(Path(export_path) / f"state_matrix_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")

        export_file = Path(export_path)
        export_file.parent.mkdir(parents=True, exist_ok=True)

        # Import lazily to avoid dashboard startup overhead.
        import pandas as pd

        frame = pd.DataFrame(rows)

        # ===== Clean 25-Column State Matrix Layout (5 Sections) =====
        ordered_columns = [
            # Section A: Identity (3 columns)
            "smart_id",
            "file_name",
            "file_path",
            # Section B: Taxonomy (12 columns)
            "metadata_level_code",
            "record_class_name",
            "record_category_name_functional",
            "record_category_name_transactional",
            "record_type_code",
            "business_unit_name",
            "sub_business_unit_name",
            "iso_country_code",
            "record_format_name",
            "original_record_location_type_name",
            "data_classification_name",
            "divestiture_deal_name",
            # Section C: Extracted Intelligence (6 columns)
            "key_names",
            "amount_found",
            "important_dates",
            "location_mentioned",
            "dynamic_subtags",
            # Section D: File Metadata (4 columns)
            "file_type",
            "file_size",
            "updated_at",
            "tag_confidence",
            # Section E: Accuracy Metrics (8 columns)
            "pipeline_type",
            "extraction_accuracy",
            "enhanced_accuracy",
            "approval_status",
            "text_area_pct",
            "non_text_area_pct",
            "preprocessing_gain_pct",
            "accuracy_loss_json",
        ]

        for col in ordered_columns:
            if col not in frame.columns:
                frame[col] = ""
        for col in ordered_columns:
            frame[col] = frame[col].fillna("")
            frame[col] = frame[col].apply(
                lambda v: "" if (pd.isna(v) or str(v).strip() == "" or str(v).strip().lower() in {"nan", "none"}) else v
            )

        frame = frame[ordered_columns]

        # Clean header rename map
        rename_map = {
            "smart_id": "Document ID",
            "file_name": "File Name",
            "file_path": "File Path",
            "metadata_level_code": "Metadata Level",
            "record_class_name": "Record Class",
            "record_category_name_functional": "Record Category (Functional)",
            "record_category_name_transactional": "Record Category (Transactional)",
            "record_type_code": "Record Type Code",
            "business_unit_name": "Business Unit",
            "sub_business_unit_name": "Sub Business Unit",
            "iso_country_code": "ISO Country Code",
            "record_format_name": "Record Format",
            "original_record_location_type_name": "Original Location Type",
            "data_classification_name": "Data Classification",
            "divestiture_deal_name": "Divestiture Deal Name",
            "key_names": "Key Names",
            "amount_found": "Amount Found",
            "important_dates": "Important Dates",
            "location_mentioned": "Locations Mentioned",
            "dynamic_subtags": "Dynamic Subtags",
            "file_type": "File Type",
            "file_size": "File Size",
            "updated_at": "Processed On",
            "tag_confidence": "Confidence Score",
            "pipeline_type": "Pipeline Type",
            "extraction_accuracy": "Extraction Accuracy %",
            "enhanced_accuracy": "Enhanced Accuracy %",
            "approval_status": "Approval Status",
            "text_area_pct": "Text Area %",
            "non_text_area_pct": "Non-Text Area %",
            "preprocessing_gain_pct": "Preprocessing Gain %",
            "accuracy_loss_json": "Accuracy Loss Reason",
        }

        frame = frame.rename(columns=rename_map)

        # ===== Build 6-Sheet Guided Workbook with openpyxl =====
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---- Color definitions ----
        BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        ORANGE_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        GRAY_FILL = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1F4E79")
        SECTION_FONT = Font(name="Calibri", bold=True, size=13, color="2E75B6")
        BODY_FONT = Font(name="Calibri", size=11)
        THIN_BORDER = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # ============================================================
        # Sheet 1: README
        # ============================================================
        ws_readme = wb.active
        ws_readme.title = "\U0001F4CB README"
        ws_readme.sheet_properties.tabColor = "4472C4"

        ws_readme.merge_cells("A1:E1")
        ws_readme["A1"].value = "GECC Document Archive \u2014 State Matrix Reference Guide"
        ws_readme["A1"].font = TITLE_FONT
        ws_readme["A1"].alignment = Alignment(horizontal="center")

        ws_readme["A3"].value = "Generated On:"
        ws_readme["A3"].font = Font(bold=True)
        ws_readme["B3"].value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        ws_readme["A4"].value = "Total Documents:"
        ws_readme["A4"].font = Font(bold=True)
        ws_readme["B4"].value = len(frame)
        ws_readme["A5"].value = "Total Columns:"
        ws_readme["A5"].font = Font(bold=True)
        ws_readme["B5"].value = len(frame.columns)

        ws_readme["A7"].value = "Workbook Structure"
        ws_readme["A7"].font = SECTION_FONT
        sheet_info = [
            ("\U0001F4CB README", "This guide \u2014 explains the workbook and how to use it"),
            ("\U0001F4CA Data Summary", "Statistical overview of all processed documents"),
            ("\U0001F50D Search Guide", "How to find specific documents using filters and search"),
            ("\U0001F4D6 Taxonomy Legend", "Valid values and definitions for every taxonomy dimension"),
            ("\U0001F4DD Column Dictionary", "Technical description of each column"),
            ("\U0001F4C4 Document Data", "The complete document data grid (use Excel filters here)"),
            ("\U0001F3AF Accuracy Report", "Extraction accuracy statistics, distribution, and analysis"),
        ]
        for i, (name, desc) in enumerate(sheet_info, 8):
            ws_readme[f"A{i}"].value = name
            ws_readme[f"A{i}"].font = Font(bold=True)
            ws_readme[f"B{i}"].value = desc

        ws_readme["A15"].value = "How to Use This Workbook"
        ws_readme["A15"].font = SECTION_FONT
        steps = [
            "Step 1: Go to '\U0001F4CA Data Summary' to see what's in this dataset at a glance",
            "Step 2: Go to '\U0001F50D Search Guide' to learn how to find specific documents",
            "Step 3: Go to '\U0001F4D6 Taxonomy Legend' to understand what each code/value means",
            "Step 4: Use the '\U0001F4C4 Document Data' sheet with Excel AutoFilters to drill into individual files",
        ]
        for i, step in enumerate(steps, 16):
            ws_readme[f"A{i}"].value = step

        ws_readme.column_dimensions["A"].width = 30
        ws_readme.column_dimensions["B"].width = 70

        # ============================================================
        # Sheet 2: Data Summary
        # ============================================================
        ws_summary = wb.create_sheet("\U0001F4CA Data Summary")
        ws_summary.sheet_properties.tabColor = "548235"

        def _write_summary_block(ws, start_row, title, series):
            ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
            ws.cell(row=start_row + 1, column=1, value="Value").font = Font(bold=True)
            ws.cell(row=start_row + 1, column=2, value="Count").font = Font(bold=True)
            ws.cell(row=start_row + 1, column=3, value="Percentage").font = Font(bold=True)
            total = series.sum() if hasattr(series, 'sum') else sum(series.values())
            r = start_row + 2
            items = series.items() if hasattr(series, 'items') else series.items()
            for val, cnt in items:
                ws.cell(row=r, column=1, value=str(val))
                ws.cell(row=r, column=2, value=int(cnt))
                pct = round(cnt / max(total, 1) * 100, 1)
                ws.cell(row=r, column=3, value=f"{pct}%")
                r += 1
            return r + 1  # next block start

        ws_summary.cell(row=1, column=1, value="Data Summary Dashboard").font = TITLE_FONT
        ws_summary.cell(row=2, column=1, value=f"Total Documents: {len(frame)}").font = Font(bold=True, size=12)

        row_cursor = 4
        for col_name in ["Record Category (Functional)", "Record Category (Transactional)", "Business Unit", "Record Type Code",
                         "ISO Country Code", "Data Classification", "File Type",
                         "Record Format", "Original Location Type", "Sub Business Unit", "Divestiture Deal Name"]:
            if col_name in frame.columns:
                vc = frame[col_name].astype(str).replace("", "(blank)").value_counts().sort_index()
                if len(vc) > 0:
                    row_cursor = _write_summary_block(ws_summary, row_cursor, f"By {col_name}", vc)

        # NLP Coverage
        ws_summary.cell(row=row_cursor, column=1, value="NLP Extraction Coverage").font = SECTION_FONT
        row_cursor += 1
        ws_summary.cell(row=row_cursor, column=1, value="Field").font = Font(bold=True)
        ws_summary.cell(row=row_cursor, column=2, value="Fill Rate").font = Font(bold=True)
        ws_summary.cell(row=row_cursor, column=3, value="Filled / Total").font = Font(bold=True)
        row_cursor += 1
        for nlp_col in ["Key Names", "Amount Found", "Important Dates", "Locations Mentioned", "Dynamic Subtags"]:
            if nlp_col in frame.columns:
                filled = frame[nlp_col].astype(str).replace("", pd.NA).dropna().shape[0]
                rate = round(filled / max(len(frame), 1) * 100, 1)
                ws_summary.cell(row=row_cursor, column=1, value=nlp_col)
                ws_summary.cell(row=row_cursor, column=2, value=f"{rate}%")
                ws_summary.cell(row=row_cursor, column=3, value=f"{filled} / {len(frame)}")
                row_cursor += 1

        # Confidence stats
        row_cursor += 1
        ws_summary.cell(row=row_cursor, column=1, value="Confidence Score Statistics").font = SECTION_FONT
        row_cursor += 1
        if "Confidence Score" in frame.columns:
            conf_vals = pd.to_numeric(frame["Confidence Score"], errors="coerce").dropna()
            if len(conf_vals) > 0:
                for stat_name, stat_val in [("Average", conf_vals.mean()), ("Min", conf_vals.min()),
                                             ("Max", conf_vals.max()), ("Std Dev", conf_vals.std())]:
                    ws_summary.cell(row=row_cursor, column=1, value=stat_name)
                    ws_summary.cell(row=row_cursor, column=2, value=round(float(stat_val), 4))
                    row_cursor += 1

        ws_summary.column_dimensions["A"].width = 40
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 20

        # ============================================================
        # Sheet 3: Search Guide
        # ============================================================
        ws_search = wb.create_sheet("\U0001F50D Search Guide")
        ws_search.sheet_properties.tabColor = "BF8F00"

        ws_search.cell(row=1, column=1, value="Quick Search Recipes").font = TITLE_FONT
        ws_search.cell(row=2, column=1, value="Use these recipes to find specific documents in the Document Data sheet using Excel AutoFilters.")

        headers = ["I Want To Find...", "Method", "What To Do", "Which Column"]
        for c, h in enumerate(headers, 1):
            cell = ws_search.cell(row=4, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = GREEN_FILL
            cell.alignment = Alignment(horizontal="center")

        recipes = [
            ("All Treasury documents", "Excel Filter", "Filter 'Business Unit' = Treasury", "Business Unit"),
            ("All contracts and agreements", "Excel Filter", "Filter 'Record Type Code' = LEG120", "Record Type Code"),
            ("Confidential files only", "Excel Filter", "Filter 'Data Classification' = GE Confidential", "Data Classification"),
            ("Documents mentioning a country", "Excel Filter", "Filter 'ISO Country Code' = USA (or other code)", "ISO Country Code"),
            ("All audit-related documents", "Excel Filter", "Filter 'Record Type Code' = AUD110", "Record Type Code"),
            ("HR / employee records", "Excel Filter", "Filter 'Record Type Code' starts with HR", "Record Type Code"),
            ("Documents with monetary amounts", "Excel Filter", "Filter 'Amount Found' is not blank", "Amount Found"),
            ("Tax-related documents", "Excel Filter", "Filter 'Record Type Code' = TAX140", "Record Type Code"),
            ("All PDF files", "Excel Filter", "Filter 'File Type' = pdf", "File Type"),
            ("Documents from a specific date", "Excel Filter", "Filter 'Important Dates' contains the date", "Important Dates"),
            ("Docs tagged with 'compliance'", "Excel Filter", "Filter 'Dynamic Subtags' contains compliance", "Dynamic Subtags"),
            ("Low-confidence classifications", "Excel Sort", "Sort 'Confidence Score' ascending", "Confidence Score"),
            ("Documents by file type", "Excel Filter", "Filter 'File Type' dropdown", "File Type"),
            ("Risk management documents", "Excel Filter", "Filter 'Record Category (Functional)' = Risk Management", "Record Category (Functional)"),
            ("Large files", "Excel Sort", "Sort 'File Size' descending", "File Size"),
        ]
        for r, (want, method, todo, col) in enumerate(recipes, 5):
            ws_search.cell(row=r, column=1, value=want)
            ws_search.cell(row=r, column=2, value=method)
            ws_search.cell(row=r, column=3, value=todo)
            ws_search.cell(row=r, column=4, value=col)

        r = len(recipes) + 7
        ws_search.cell(row=r, column=1, value="Dashboard Search (Streamlit UI)").font = SECTION_FONT
        r += 1
        ws_search.cell(row=r, column=1, value="1. Open the dashboard at http://localhost:8501")
        r += 1
        ws_search.cell(row=r, column=1, value="2. Enter keywords in the search box (e.g., 'treasury cash management')")
        r += 1
        ws_search.cell(row=r, column=1, value="3. Use Advanced Filters to narrow by File Type, Category, Department, Business Unit")
        r += 1
        ws_search.cell(row=r, column=1, value="API Search: GET /search?q=treasury&size=20")

        ws_search.column_dimensions["A"].width = 40
        ws_search.column_dimensions["B"].width = 15
        ws_search.column_dimensions["C"].width = 55
        ws_search.column_dimensions["D"].width = 20

        # ============================================================
        # Sheet 4: Taxonomy Legend
        # ============================================================
        ws_tax = wb.create_sheet("\U0001F4D6 Taxonomy Legend")
        ws_tax.sheet_properties.tabColor = "7030A0"

        ws_tax.cell(row=1, column=1, value="Taxonomy Dimension Reference").font = TITLE_FONT

        # Record Type Code descriptions
        _code_descriptions = {
            "FIN210": "General Ledger / Chart of Accounts",
            "FIN200": "Journal Entries / Postings",
            "FIN100": "Accounts Payable",
            "FIN110": "Accounts Receivable",
            "FIN160": "Financial Reports / Statements",
            "FIN180": "Treasury / Cash Management",
            "TAX140": "Tax Returns / Filings",
            "HR120": "Employee / Personnel Files",
            "HR160": "Payroll / Compensation",
            "HR180": "HR Operations / Recruitment",
            "LEG120": "Contracts / Agreements",
            "LEG160": "Legal Advice / Opinions",
            "AUD110": "Audit / SOX Controls",
            "INS120": "Insurance / Claims",
            "REG140": "Regulatory / Compliance",
            "RSK120": "Risk Assessment / Management",
            "CAD100": "Reports / Analysis",
            "CAD110": "Budget / Forecasting",
            "CAD120": "Meeting Minutes / Agendas",
            "CAD150": "Correspondence / Memos",
            "CAD180": "Policies / Procedures",
            "CAD190": "Project Plans / Milestones",
            "CAD210": "Training Records / Materials",
            "CAD230": "Customer / Client Records",
            "SOU110": "Vendor / Procurement",
            "MKT100": "Marketing / Campaigns",
            "TEC100": "IT Systems / Infrastructure",
            "TEC120": "Technical Specifications",
            "FM110": "Facilities / Building Management",
            "RIM100": "Records Management / Retention",
            "BDV110": "Business Development / Strategy",
        }

        tax_row = 3
        taxonomy_dims = [
            ("Record Category (Functional)", "The business function domain this document belongs to"),
            ("Record Category (Transactional)", "Transactional document sub-category"),
            ("Record Type Code", "Specific document classification code within its category"),
            ("Business Unit", "The organizational unit that owns/manages this record"),
            ("Sub Business Unit", "Department subdivision within the business unit"),
            ("ISO Country Code", "Country jurisdiction referenced in the document"),
            ("Data Classification", "Sensitivity level determining handling requirements"),
            ("Record Format", "Whether the record is Electronic or Physical (scanned)"),
            ("Original Location Type", "Where the record was originally stored"),
            ("Metadata Level", "Granularity of metadata: File or Folder"),
            ("Record Class", "High-level classification: Functional or Transactional"),
            ("Divestiture Deal Name", "Target corporate divestiture or deal identifier if mentioned"),
        ]

        for dim_name, dim_desc in taxonomy_dims:
            ws_tax.cell(row=tax_row, column=1, value=dim_name).font = SECTION_FONT
            ws_tax.cell(row=tax_row, column=2, value=dim_desc)
            tax_row += 1

            # Write valid values
            ws_tax.cell(row=tax_row, column=1, value="Value").font = Font(bold=True)
            ws_tax.cell(row=tax_row, column=2, value="Description").font = Font(bold=True)
            ws_tax.cell(row=tax_row, column=3, value="Count in Dataset").font = Font(bold=True)
            tax_row += 1

            if dim_name in frame.columns:
                vc = frame[dim_name].astype(str).replace("", "(blank)").value_counts().sort_index()
                for val, cnt in vc.items():
                    ws_tax.cell(row=tax_row, column=1, value=str(val))
                    # Add description for Record Type Codes
                    if dim_name == "Record Type Code" and val in _code_descriptions:
                        ws_tax.cell(row=tax_row, column=2, value=_code_descriptions[val])
                    elif dim_name == "Data Classification":
                        _dc_desc = {
                            "Public": "No restrictions \u2014 may be shared externally",
                            "GE Internal": "For internal use only \u2014 not for external distribution",
                            "GE Confidential": "Restricted access \u2014 need-to-know basis",
                            "GE Restricted": "Highly restricted \u2014 limited authorized personnel",
                            "GE Confidential with SPII": "Contains sensitive personal data \u2014 strictest controls",
                        }
                        ws_tax.cell(row=tax_row, column=2, value=_dc_desc.get(val, ""))
                    ws_tax.cell(row=tax_row, column=3, value=int(cnt))
                    tax_row += 1
            tax_row += 1

        ws_tax.column_dimensions["A"].width = 35
        ws_tax.column_dimensions["B"].width = 60
        ws_tax.column_dimensions["C"].width = 18

        # ============================================================
        # Sheet 5: Column Dictionary
        # ============================================================
        ws_dict = wb.create_sheet("\U0001F4DD Column Dictionary")
        ws_dict.sheet_properties.tabColor = "C00000"

        ws_dict.cell(row=1, column=1, value="Column Dictionary \u2014 23 Columns").font = TITLE_FONT

        dict_headers = ["#", "Column Name", "Section", "Data Type", "Always Filled?", "Fill Rate", "Description"]
        for c, h in enumerate(dict_headers, 1):
            cell = ws_dict.cell(row=3, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        col_defs = [
            ("Document ID", "Identity", "String", True, "Unique document identifier (e.g., DOC-20260525-B2E3)"),
            ("File Name", "Identity", "String", True, "Original filename with extension"),
            ("File Path", "Identity", "String", True, "Full filesystem path to the source document"),
            ("Metadata Level", "Taxonomy", "String", True, "Granularity: always 'File' for file-level processing"),
            ("Record Class", "Taxonomy", "String", True, "Functional or Undefined, based on whether a category was matched"),
            ("Record Category (Functional)", "Taxonomy", "String", True, "Business domain derived from Record Type Code (e.g., Legal, HR, Finance)"),
            ("Record Category (Transactional)", "Taxonomy", "String", False, "Transactional document sub-category derived from content signals"),
            ("Record Type Code", "Taxonomy", "String", True, "Specific document type code from content keyword matching"),
            ("Business Unit", "Taxonomy", "String", True, "Organizational unit derived from content signals"),
            ("Sub Business Unit", "Taxonomy", "String", False, "Sub-division within parent BU, from registry matching"),
            ("ISO Country Code", "Taxonomy", "String", True, "Country jurisdiction from content analysis (ISO-3)"),
            ("Record Format", "Taxonomy", "String", True, "Electronic or Physical (scanned/OCR-only documents)"),
            ("Original Location Type", "Taxonomy", "String", True, "Storage origin: Shared Drive, Email, SharePoint, etc."),
            ("Data Classification", "Taxonomy", "String", True, "Sensitivity level: Public to GE Confidential with SPII"),
            ("Divestiture Deal Name", "Taxonomy", "String", False, "Target corporate divestiture or deal identifier if mentioned in content"),
            ("Key Names", "Intelligence", "String", False, "Important proper nouns and entities extracted from content"),
            ("Amount Found", "Intelligence", "String", False, "Monetary amounts with currency symbols found in text"),
            ("Important Dates", "Intelligence", "String", False, "Calendar dates mentioned in the document"),
            ("Locations Mentioned", "Intelligence", "String", False, "Geographic locations referenced in document content"),
            ("Dynamic Subtags", "Intelligence", "String", False, "Top content-relevant keywords from taxonomy matching"),
            ("File Type", "Metadata", "String", True, "File extension (pdf, docx, xlsx, etc.)"),
            ("File Size", "Metadata", "Integer", True, "File size in bytes"),
            ("Processed On", "Metadata", "DateTime", True, "UTC timestamp of when the document was last processed"),
            ("Confidence Score", "Metadata", "Float", True, "Average classification confidence (0.0 to 1.0)"),
            ("Pipeline Type", "Accuracy", "String", True, "How the document was processed: 'text_extraction' (digital formats) or 'ocr' (scanned/image)"),
            ("Extraction Accuracy %", "Accuracy", "Number", True, "0-100 percentage of how much text content was successfully extracted from the document"),
            ("Enhanced Accuracy %", "Accuracy", "Number", True, "Accuracy score adjusted and boosted after human audit accepts visual elements"),
            ("Approval Status", "Accuracy", "String", True, "Status of human review: 'Full Baseline', 'Pending Review', or 'Approved'"),
            ("Text Area %", "Accuracy", "Number", True, "Percentage of the document's page area that contains readable text"),
            ("Non-Text Area %", "Accuracy", "Number", True, "Percentage of the page consumed by non-text elements (images, logos, signatures, stamps)"),
            ("Preprocessing Gain %", "Accuracy", "Number", True, "For OCR documents: how much preprocessing improved character recognition (0% for digital formats)"),
            ("Accuracy Loss Reason", "Accuracy", "String", False, "Human-readable breakdown of what reduced extraction accuracy (e.g., 'Logos: 15%; Signatures: 5%')"),
        ]

        section_fill_map = {
            "Identity": BLUE_FILL,
            "Taxonomy": GREEN_FILL,
            "Intelligence": ORANGE_FILL,
            "Metadata": GRAY_FILL,
            "Accuracy": PatternFill(start_color="008080", end_color="008080", fill_type="solid"),
        }

        for i, (name, section, dtype, always, desc) in enumerate(col_defs, 1):
            r = i + 3
            ws_dict.cell(row=r, column=1, value=i)
            ws_dict.cell(row=r, column=2, value=name).font = Font(bold=True)
            ws_dict.cell(row=r, column=3, value=section)
            ws_dict.cell(row=r, column=4, value=dtype)
            ws_dict.cell(row=r, column=5, value="Yes" if always else "No")
            # Compute fill rate
            if name in frame.columns:
                filled = frame[name].astype(str).replace("", pd.NA).dropna().shape[0]
                rate = round(filled / max(len(frame), 1) * 100, 1)
            else:
                rate = 0.0
            ws_dict.cell(row=r, column=6, value=f"{rate}%")
            ws_dict.cell(row=r, column=7, value=desc)
            # Apply section-based row fill
            if section in section_fill_map:
                fill = section_fill_map[section]
                for col_idx in range(1, 8):
                    ws_dict.cell(row=r, column=col_idx).fill = fill
                    ws_dict.cell(row=r, column=col_idx).font = Font(bold=True, color="FFFFFF")

        for c, w in enumerate([5, 22, 14, 10, 14, 10, 65], 1):
            ws_dict.column_dimensions[get_column_letter(c)].width = w

        # ============================================================
        # Sheet 6: Document Data
        # ============================================================
        ws_data = wb.create_sheet("\U0001F4C4 Document Data")
        ws_data.sheet_properties.tabColor = "2E75B6"

        # Section color mapping: column index -> fill
        section_fills = {}
        for c in range(1, 4):   # Identity: cols 1-3
            section_fills[c] = BLUE_FILL
        for c in range(4, 16):  # Taxonomy: cols 4-15
            section_fills[c] = GREEN_FILL
        for c in range(16, 22): # Intelligence: cols 16-21
            section_fills[c] = ORANGE_FILL
        for c in range(22, 26): # Metadata: cols 22-25
            section_fills[c] = GRAY_FILL
        TEAL_FILL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        for c in range(26, 34): # Accuracy: cols 26-33
            section_fills[c] = TEAL_FILL


        # Write header row
        for c, col_name in enumerate(frame.columns, 1):
            cell = ws_data.cell(row=1, column=c, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = section_fills.get(c, GRAY_FILL)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Define soft green highlights for approved/enhanced metrics
        GREEN_HIGHLIGHT = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        GREEN_TEXT = Font(name="Calibri", size=11, color="375623", bold=True)

        try:
            enhanced_acc_col_idx = list(frame.columns).index("Enhanced Accuracy %") + 1
            approval_status_col_idx = list(frame.columns).index("Approval Status") + 1
        except ValueError:
            enhanced_acc_col_idx = None
            approval_status_col_idx = None

        # Write data rows
        for r_idx, (_, row_data) in enumerate(frame.iterrows(), 2):
            is_approved = False
            if approval_status_col_idx:
                status_val = row_data.iloc[approval_status_col_idx - 1]
                if str(status_val).strip() == "Approved":
                    is_approved = True

            for c_idx, val in enumerate(row_data, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

                # Accent green highlight for approved visual audits
                if is_approved and c_idx in (enhanced_acc_col_idx, approval_status_col_idx):
                    cell.fill = GREEN_HIGHLIGHT
                    cell.font = GREEN_TEXT

        # Freeze header row
        ws_data.freeze_panes = "A2"

        # Enable AutoFilter
        if len(frame) > 0:
            last_col = get_column_letter(len(frame.columns))
            ws_data.auto_filter.ref = f"A1:{last_col}{len(frame) + 1}"

        # Auto-fit column widths
        for c in range(1, len(frame.columns) + 1):
            col_letter = get_column_letter(c)
            max_len = len(str(frame.columns[c - 1]))
            for r in range(2, min(len(frame) + 2, 52)):  # sample first 50 rows
                val = ws_data.cell(row=r, column=c).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws_data.column_dimensions[col_letter].width = max(12, min(max_len + 2, 45))

        # ============================================================
        # Sheet 7: Accuracy Report (with Charts & Visualizations)
        # ============================================================
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.series import DataPoint, SeriesLabel
        from openpyxl.chart.label import DataLabelList

        ws_acc = wb.create_sheet("\U0001F3AF Accuracy Report")
        ws_acc.sheet_properties.tabColor = "008080"
        TEAL_FILL_HDR = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        METRIC_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        VALUE_FONT = Font(name="Calibri", size=11, color="333333")

        ws_acc.cell(row=1, column=1, value="Document Extraction Accuracy Report").font = TITLE_FONT
        ws_acc.cell(row=2, column=1, value=f"Total Documents: {len(frame)}").font = Font(bold=True, size=12)
        ws_acc.cell(row=3, column=1,
                    value="This report summarizes how accurately text was extracted from each document. "
                          "It shows what percentage of each document the system could read, and explains "
                          "what caused any accuracy loss (e.g., logos, images, signatures, noise)."
        ).font = Font(italic=True, size=10, color="666666")

        acc_col_name = "Extraction Accuracy %"
        pipe_col_name = "Pipeline Type"
        ftype_col_name = "File Type"
        prep_col_name = "Preprocessing Gain %"
        text_col_name = "Text Area %"
        nontext_col_name = "Non-Text Area %"

        row_c = 5

        # ---- Section 1: Overall Accuracy Statistics ----
        ws_acc.cell(row=row_c, column=1, value="1. Overall Accuracy Statistics").font = SECTION_FONT
        row_c += 1
        if acc_col_name in frame.columns:
            acc_vals = pd.to_numeric(frame[acc_col_name], errors="coerce").dropna()
            if len(acc_vals) > 0:
                stats = [
                    ("Average Accuracy", f"{acc_vals.mean():.2f}%", "The mean extraction accuracy across all documents"),
                    ("Median Accuracy", f"{acc_vals.median():.2f}%", "50th percentile — half the documents score above this"),
                    ("Min Accuracy", f"{acc_vals.min():.2f}%", "The lowest-scoring document's accuracy"),
                    ("Max Accuracy", f"{acc_vals.max():.2f}%", "The highest-scoring document's accuracy"),
                    ("Std Deviation", f"{acc_vals.std():.2f}", "Spread of accuracy scores — lower means more consistent"),
                    ("Documents Measured", str(int(len(acc_vals))), "Total documents with accuracy data"),
                ]
                for hdr_c, hdr in enumerate(["Metric", "Value", "What This Means"], 1):
                    c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                    c.font = WHITE_FONT
                    c.fill = TEAL_FILL_HDR
                row_c += 1
                for label, val, desc in stats:
                    ws_acc.cell(row=row_c, column=1, value=label).font = METRIC_FONT
                    ws_acc.cell(row=row_c, column=2, value=val).font = VALUE_FONT
                    ws_acc.cell(row=row_c, column=3, value=desc).font = Font(size=10, color="555555")
                    row_c += 1
            else:
                ws_acc.cell(row=row_c, column=1, value="No accuracy data available yet")
                row_c += 1

        # ---- Section 2: Accuracy by Pipeline Type ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="2. Accuracy by Pipeline Type").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="'text_extraction' = digital formats (DOCX, XLSX, TXT, etc.) | 'ocr' = scanned/image documents"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2
        pipe_table_start = row_c
        if pipe_col_name in frame.columns and acc_col_name in frame.columns:
            for hdr_c, hdr in enumerate(["Pipeline", "Avg Accuracy %", "Doc Count"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1
            grouped = frame.groupby(pipe_col_name)[acc_col_name]
            for pipe_name, group in grouped:
                vals = pd.to_numeric(group, errors="coerce").dropna()
                if len(vals) > 0:
                    ws_acc.cell(row=row_c, column=1, value=str(pipe_name) if str(pipe_name).strip() else "(not set)")
                    ws_acc.cell(row=row_c, column=2, value=round(float(vals.mean()), 2))
                    ws_acc.cell(row=row_c, column=3, value=len(vals))
                    row_c += 1

        # ---- Section 3: Accuracy by File Type (with chart) ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="3. Accuracy by File Type").font = SECTION_FONT
        row_c += 1
        ftype_table_start = row_c
        ftype_rows = 0
        if ftype_col_name in frame.columns and acc_col_name in frame.columns:
            for hdr_c, hdr in enumerate(["File Type", "Avg Accuracy %", "Doc Count", "Min %", "Max %"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1
            grouped = frame.groupby(ftype_col_name)[acc_col_name]
            for ft, group in sorted(grouped, key=lambda x: x[0]):
                vals = pd.to_numeric(group, errors="coerce").dropna()
                if len(vals) > 0:
                    ws_acc.cell(row=row_c, column=1, value=str(ft).upper() if str(ft).strip() else "(unknown)")
                    ws_acc.cell(row=row_c, column=2, value=round(float(vals.mean()), 2))
                    ws_acc.cell(row=row_c, column=3, value=len(vals))
                    ws_acc.cell(row=row_c, column=4, value=round(float(vals.min()), 2))
                    ws_acc.cell(row=row_c, column=5, value=round(float(vals.max()), 2))
                    row_c += 1
                    ftype_rows += 1


        # ---- Section 4: Preprocessing Impact ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="4. Preprocessing Impact (OCR Documents)").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="Shows how much image preprocessing (deskew, threshold, contrast) improved text recognition"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2
        if prep_col_name in frame.columns:
            prep_vals = pd.to_numeric(frame[prep_col_name], errors="coerce").dropna()
            improved = prep_vals[prep_vals > 0]
            for label, val in [
                ("Average Preprocessing Gain", f"{prep_vals.mean():.2f}%" if len(prep_vals) > 0 else "0.00%"),
                ("Documents Where Preprocessing Helped", str(len(improved))),
                ("Best Improvement Achieved", f"{improved.max():.2f}%" if len(improved) > 0 else "N/A"),
                ("Documents With No Gain (digital formats)", str(int((prep_vals == 0).sum()))),
            ]:
                ws_acc.cell(row=row_c, column=1, value=label).font = METRIC_FONT
                ws_acc.cell(row=row_c, column=2, value=val).font = VALUE_FONT
                row_c += 1

        # ---- Section 5: Accuracy Distribution (with Pie Chart) ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="5. Accuracy Distribution").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="How many documents fall into each quality tier"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2
        dist_table_start = row_c
        dist_rows = 0
        if acc_col_name in frame.columns:
            acc_numeric = pd.to_numeric(frame[acc_col_name], errors="coerce").dropna()
            bins = [
                ("90-100% (Excellent)", int((acc_numeric >= 90).sum()), "00B050"),
                ("80-90% (Good)", int(((acc_numeric >= 80) & (acc_numeric < 90)).sum()), "92D050"),
                ("70-80% (Fair)", int(((acc_numeric >= 70) & (acc_numeric < 80)).sum()), "FFC000"),
                ("< 70% (Low)", int((acc_numeric < 70).sum()), "FF4444"),
            ]
            for hdr_c, hdr in enumerate(["Quality Tier", "Document Count", "Percentage"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1
            for label, count, color in bins:
                pct = round(count / max(len(acc_numeric), 1) * 100, 1)
                ws_acc.cell(row=row_c, column=1, value=label)
                ws_acc.cell(row=row_c, column=2, value=count)
                ws_acc.cell(row=row_c, column=3, value=f"{pct}%")
                row_c += 1
                dist_rows += 1



        # ---- Section 6: Text Area vs Non-Text Area by File Type (Stacked Bar) ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="6. Page Composition: Text vs Non-Text by File Type").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="Shows the average page layout composition — how much of each file type is readable text vs non-text (images, logos, signatures, stamps)"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2
        comp_table_start = row_c
        comp_rows = 0
        if text_col_name in frame.columns and nontext_col_name in frame.columns and ftype_col_name in frame.columns:
            for hdr_c, hdr in enumerate(["File Type", "Avg Text Area %", "Avg Non-Text Area %"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1
            for ft in sorted(frame[ftype_col_name].unique()):
                ft_str = str(ft).strip()
                if not ft_str:
                    continue
                mask = frame[ftype_col_name] == ft
                txt_vals = pd.to_numeric(frame.loc[mask, text_col_name], errors="coerce").dropna()
                ntxt_vals = pd.to_numeric(frame.loc[mask, nontext_col_name], errors="coerce").dropna()
                if len(txt_vals) > 0:
                    ws_acc.cell(row=row_c, column=1, value=ft_str.upper())
                    ws_acc.cell(row=row_c, column=2, value=round(float(txt_vals.mean()), 2))
                    ws_acc.cell(row=row_c, column=3, value=round(float(ntxt_vals.mean()), 2) if len(ntxt_vals) > 0 else 0)
                    row_c += 1
                    comp_rows += 1



        # ---- Section 7: Accuracy Loss Reason Analysis ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="7. Accuracy Loss Reason Analysis").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="Aggregated view of WHY accuracy was lost across all documents — which factors consumed the most page area"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2

        # Aggregate loss reasons across all documents
        loss_col_name = "Accuracy Loss Reason"
        # Only aggregate keys that represent actual loss factors (percentages/areas)
        # Skip metadata keys like cell counts, page counts, format names
        LOSS_KEYS = {
            # OCR loss breakdown
            "text_read_pct", "unreadable_text_pct", "logos_images_pct",
            "signatures_pct", "stamps_seals_pct", "noise_artifacts_pct",
            "whitespace_margins_pct",
            # HTML/XML/JSON/CSV structural overhead
            "html_tags_pct", "syntax_overhead_pct", "image_area_pct",
            "image_area_estimate_pct", "empty_cell_pct",
            "embedded_objects_estimate",
        }
        LOSS_KEY_LABELS = {
            "html_tags_pct": "HTML Tag Overhead",
            "syntax_overhead_pct": "Format Syntax Overhead",
            "image_area_pct": "Embedded Images/Logos",
            "image_area_estimate_pct": "Estimated Image Area",
            "empty_cell_pct": "Empty Cells (Spreadsheets)",
            "embedded_objects_estimate": "Embedded Objects",
            "text_read_pct": "Text Successfully Read",
            "unreadable_text_pct": "Unreadable/Lost Text",
            "logos_images_pct": "Logos & Images",
            "signatures_pct": "Signatures",
            "stamps_seals_pct": "Stamps & Seals",
            "noise_artifacts_pct": "Noise & Artifacts",
            "whitespace_margins_pct": "Whitespace & Margins",
        }
        loss_categories = {}
        if loss_col_name in frame.columns:
            for val in frame[loss_col_name]:
                if not val or str(val).strip() in {"", "nan", "None"}:
                    continue
                try:
                    d = json.loads(str(val))
                    if isinstance(d, dict):
                        for k, v in d.items():
                            if k in LOSS_KEYS and isinstance(v, (int, float)) and v > 0:
                                label = LOSS_KEY_LABELS.get(k, k.replace("_", " ").title())
                                loss_categories[label] = loss_categories.get(label, 0) + float(v)
                except Exception:
                    pass

        # Show aggregated loss reasons as a table
        loss_table_start = row_c
        loss_rows = 0
        if loss_categories:
            # Sort by total impact (descending)
            sorted_cats = sorted(loss_categories.items(), key=lambda x: -x[1])
            total_impact = sum(v for _, v in sorted_cats)

            for hdr_c, hdr in enumerate(["Loss Factor", "Cumulative % Across All Docs", "Share of All Losses", "What This Means"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1

            # Descriptions for each loss category
            LOSS_DESCRIPTIONS = {
                "HTML Tag Overhead": "Percentage of HTML files consumed by markup tags (not readable content)",
                "Format Syntax Overhead": "Percentage consumed by format-specific syntax (JSON brackets, XML tags, CSV delimiters)",
                "Embedded Images/Logos": "Page area occupied by embedded images and company logos",
                "Estimated Image Area": "Estimated area consumed by referenced images in the document",
                "Empty Cells (Spreadsheets)": "Percentage of spreadsheet cells that are empty and contain no data",
                "Embedded Objects": "Non-text objects embedded in the document (charts, shapes)",
                "Text Successfully Read": "Page area where text was successfully extracted by the OCR engine",
                "Unreadable/Lost Text": "Page area containing text that the OCR engine could not read",
                "Logos & Images": "Page area consumed by logos, photos, and graphical elements",
                "Signatures": "Page area consumed by handwritten or digital signatures",
                "Stamps & Seals": "Page area consumed by official stamps, seals, or watermarks",
                "Noise & Artifacts": "Page area with scanning noise, smudges, or image artifacts",
                "Whitespace & Margins": "Page area that is blank space (margins, line spacing, gaps)",
            }

            for cat, total_val in sorted_cats[:15]:  # Top 15 reasons
                share_pct = round(total_val / max(total_impact, 1) * 100, 1)
                ws_acc.cell(row=row_c, column=1, value=cat)
                ws_acc.cell(row=row_c, column=2, value=round(total_val, 2))
                ws_acc.cell(row=row_c, column=3, value=f"{share_pct}%")
                ws_acc.cell(row=row_c, column=4, value=LOSS_DESCRIPTIONS.get(cat, "")).font = Font(size=9, color="666666")
                row_c += 1
                loss_rows += 1



        # ---- Section 8: Bottom 10 Lowest Accuracy Documents ----
        row_c += 2
        ws_acc.cell(row=row_c, column=1, value="8. Bottom 10 Lowest Accuracy Documents").font = SECTION_FONT
        ws_acc.cell(row=row_c + 1, column=1,
                    value="These documents had the worst extraction results — investigate their loss reasons for improvement opportunities"
        ).font = Font(italic=True, size=9, color="888888")
        row_c += 2
        bottom10_start = row_c
        bottom10_rows = 0
        if acc_col_name in frame.columns and "File Name" in frame.columns:
            frame_with_acc = frame.copy()
            frame_with_acc["_acc_num"] = pd.to_numeric(frame_with_acc[acc_col_name], errors="coerce")
            bottom10 = frame_with_acc.dropna(subset=["_acc_num"]).nsmallest(10, "_acc_num")
            for hdr_c, hdr in enumerate(["File Name", "Accuracy %", "Pipeline", "File Type", "Loss Reason Summary"], 1):
                c = ws_acc.cell(row=row_c, column=hdr_c, value=hdr)
                c.font = WHITE_FONT
                c.fill = TEAL_FILL_HDR
            row_c += 1
            for _, brow in bottom10.iterrows():
                ws_acc.cell(row=row_c, column=1, value=str(brow.get("File Name", "")))
                ws_acc.cell(row=row_c, column=2, value=brow.get("_acc_num", 0))
                ws_acc.cell(row=row_c, column=3, value=str(brow.get(pipe_col_name, "")))
                ws_acc.cell(row=row_c, column=4, value=str(brow.get(ftype_col_name, "")))
                # Format loss reason for bottom 10
                loss_raw = brow.get(loss_col_name, "")
                try:
                    loss_d = json.loads(str(loss_raw)) if loss_raw else {}
                    loss_parts = []
                    for lk, lv in (loss_d.items() if isinstance(loss_d, dict) else []):
                        fl = LOSS_KEY_LABELS.get(lk, lk.replace("_", " ").title())
                        if isinstance(lv, float):
                            loss_parts.append(f"{fl}: {lv:.1f}%")
                        elif isinstance(lv, bool):
                            loss_parts.append(f"{fl}: {'Yes' if lv else 'No'}")
                        elif isinstance(lv, int):
                            loss_parts.append(f"{fl}: {lv}")
                        elif isinstance(lv, str) and lk == "format":
                            loss_parts.append(f"Format: {lv.upper()}")
                        else:
                            loss_parts.append(f"{fl}: {lv}")
                    formatted_loss = "; ".join(loss_parts[:5]) if loss_parts else ""
                except Exception:
                    formatted_loss = str(loss_raw)[:120] if loss_raw else ""
                ws_acc.cell(row=row_c, column=5, value=formatted_loss)
                row_c += 1
                bottom10_rows += 1



        # ---- Column widths ----
        ws_acc.column_dimensions["A"].width = 45
        ws_acc.column_dimensions["B"].width = 22
        ws_acc.column_dimensions["C"].width = 18
        ws_acc.column_dimensions["D"].width = 15
        ws_acc.column_dimensions["E"].width = 55


        # ============================================================
        # Format accuracy_loss_json → clear, human-readable explanations
        # in the Document Data sheet
        # ============================================================
        if loss_col_name in frame.columns:
            # Mapping of raw JSON keys to clear, user-friendly labels
            FRIENDLY_LABELS = {
                # HTML
                "html_tags_pct": "HTML Tags",
                "embedded_images": "Embedded Images",
                "image_area_estimate_pct": "Estimated Image Area",
                # JSON / XML / CSV / MD / TXT
                "syntax_overhead_pct": "Format Syntax Overhead",
                "format": "File Format",
                # DOCX
                "paragraphs": "Paragraph Count",
                "tables": "Table Count",
                "table_chars": "Characters in Tables",
                "images": "Embedded Images",
                "header_footer_chars": "Header/Footer Characters",
                "image_area_pct": "Image Area",
                # XLSX
                "total_cells": "Total Cells",
                "text_cells": "Text Cells",
                "number_cells": "Number Cells",
                "empty_cells": "Empty Cells",
                "empty_cell_pct": "Empty Cell Percentage",
                "embedded_objects_estimate": "Estimated Embedded Objects",
                # PDF
                "page_count": "Total Pages",
                "chars_per_page": "Characters Per Page",
                "text_density": "Text Density Ratio",
                "low_text_warning": "Low Text Warning",
                # OCR loss breakdown
                "text_read_pct": "Text Successfully Read",
                "unreadable_text_pct": "Unreadable Text (Lost)",
                "logos_images_pct": "Logos & Images",
                "signatures_pct": "Signatures",
                "stamps_seals_pct": "Stamps & Seals",
                "noise_artifacts_pct": "Noise & Artifacts",
                "whitespace_margins_pct": "Whitespace & Margins",
                # Error fallback
                "error": "Analysis Error",
                "method": "Analysis Method",
            }

            def _format_loss_reason(raw_val, baseline_acc):
                """Convert accuracy_loss_json to a clear, human-readable explanation."""
                try:
                    baseline_val = float(baseline_acc) if baseline_acc is not None and str(baseline_acc).strip() != "" else 100.0
                except (ValueError, TypeError):
                    baseline_val = 100.0
                
                loss_pct = 100.0 - baseline_val
                if loss_pct <= 0:
                    return "No accuracy loss detected — full extraction achieved"

                if not raw_val or str(raw_val).strip() in {"", "nan", "None", "{}"}:
                    return "No accuracy loss detected — full extraction achieved"
                try:
                    d = json.loads(str(raw_val))
                    if not isinstance(d, dict) or not d:
                        return f"Lost {loss_pct:.1f}% of content due to formatting/structure"

                    # If this is OCR/Image pipeline, format as requested: "Lost 30% of content: 15% Signatures, 10% Stamps..."
                    ocr_keys = {"unreadable_text_pct", "signatures_pct", "stamps_seals_pct", "logos_images_pct", "noise_artifacts_pct", "whitespace_margins_pct"}
                    has_ocr_keys = any(k in d for k in ocr_keys)

                    if has_ocr_keys:
                        factors = []
                        for key, label in [
                            ("unreadable_text_pct", "Unreadable Text"),
                            ("stamps_seals_pct", "Stamps & Seals"),
                            ("signatures_pct", "Signatures"),
                            ("logos_images_pct", "Logos & Images"),
                            ("noise_artifacts_pct", "Noise & Artifacts"),
                            ("whitespace_margins_pct", "Whitespace & Margins"),
                        ]:
                            val = d.get(key, 0.0)
                            if isinstance(val, (int, float)) and val > 0:
                                factors.append(f"{val:.1f}% {label}")
                        if factors:
                            return f"Lost {loss_pct:.1f}% of content: " + ", ".join(factors)
                        else:
                            return f"Lost {loss_pct:.1f}% of content due to formatting/structure"

                    # Fallback to general formatting for other document types
                    parts = []
                    for key, val in d.items():
                        if key == "snippets":  # skip raw snippet coordinates list
                            continue
                        friendly = FRIENDLY_LABELS.get(key, key.replace("_", " ").title())

                        # Format values based on type and context
                        if isinstance(val, bool):
                            parts.append(f"{friendly}: {'Yes' if val else 'No'}")
                        elif isinstance(val, float):
                            if "pct" in key or "area" in key or "overhead" in key or "density" in key:
                                parts.append(f"{friendly}: {val:.1f}%")
                            else:
                                parts.append(f"{friendly}: {val:.1f}")
                        elif isinstance(val, int):
                            parts.append(f"{friendly}: {val}")
                        elif isinstance(val, str):
                            if key == "format":
                                parts.append(f"Format: {val.upper()}")
                            else:
                                parts.append(f"{friendly}: {val}")

                    if not parts:
                        return f"Lost {loss_pct:.1f}% of content (format/structure)"

                    return f"Lost {loss_pct:.1f}%: " + "; ".join(parts)

                except (json.JSONDecodeError, TypeError, ValueError):
                    return f"Lost {loss_pct:.1f}% of content: {str(raw_val)[:100]}"

            # Apply formatting to Document Data sheet
            loss_col_idx = list(frame.columns).index(loss_col_name) + 1 if loss_col_name in frame.columns else None
            acc_col_idx = list(frame.columns).index("Extraction Accuracy %") + 1 if "Extraction Accuracy %" in frame.columns else None
            if loss_col_idx:
                for r_idx in range(2, len(frame) + 2):
                    cell = ws_data.cell(row=r_idx, column=loss_col_idx)
                    baseline_acc = 100.0
                    if acc_col_idx:
                        acc_val = ws_data.cell(row=r_idx, column=acc_col_idx).value
                        if acc_val is not None:
                            baseline_acc = acc_val
                    if cell.value:
                        cell.value = _format_loss_reason(cell.value, baseline_acc)

        # ============================================================
        # Sheet 8: 📊 Accuracy Dashboard (Matplotlib Hybrid Sheet)
        # All charts generated in-memory via Matplotlib and embedded
        # as high-resolution crisp static images.
        # ============================================================
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from io import BytesIO
        from openpyxl.drawing.image import Image as OPImage

        ws_dash = wb.create_sheet("📊 Accuracy Dashboard")
        ws_dash.sheet_properties.tabColor = "1F4E79"

        # Set column widths for clean grid layout
        for col_letter in ["A"]:
            ws_dash.column_dimensions[col_letter].width = 3
        for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
            ws_dash.column_dimensions[col_letter].width = 12

        # Set row heights for visual balance
        ws_dash.row_dimensions[1].height = 28
        ws_dash.row_dimensions[2].height = 18
        ws_dash.row_dimensions[4].height = 28
        ws_dash.row_dimensions[5].height = 16

        # ---- Row 1-2: Title ----
        ws_dash.merge_cells("B1:O1")
        title_cell = ws_dash.cell(row=1, column=2, value="Document Extraction Accuracy Dashboard")
        title_cell.font = Font(name="Segoe UI", bold=True, size=20, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_dash.merge_cells("B2:O2")
        from datetime import datetime as _dt
        tier_str = "TIER 3 (YOLOv8 + DOCTR)"
        if "accuracy_tier" in frame.columns:
            tiers = frame["accuracy_tier"].dropna().unique()
            if len(tiers) > 0:
                tier_str = " & ".join(sorted([str(t).upper() for t in tiers]))
        
        subtitle = ws_dash.cell(row=2, column=2,
            value=f"{len(frame)} documents analyzed  •  Active Accuracy Tier: {tier_str}  •  Exported {_dt.now().strftime('%Y-%m-%d %H:%M')}")
        subtitle.font = Font(name="Segoe UI", italic=True, size=10, color="555555")
        subtitle.alignment = Alignment(horizontal="center", vertical="center")

        # ---- Row 4: KPI Scorecards ----
        KPI_FILL_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        KPI_FILL_GREEN = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        KPI_FILL_RED = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        KPI_FILL_TEAL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        KPI_NUM_FONT = Font(name="Segoe UI", bold=True, size=22, color="FFFFFF")
        KPI_LABEL_FONT = Font(name="Segoe UI", bold=True, size=9, color="FFFFFF")

        enhanced_col = "Enhanced Accuracy %"
        acc_col_for_dashboard = enhanced_col if enhanced_col in frame.columns else acc_col_name

        acc_vals = pd.to_numeric(frame.get(acc_col_for_dashboard, pd.Series(dtype=float)), errors="coerce").dropna()
        kpi_data = []
        if len(acc_vals) > 0:
            kpi_data = [
                ("B", "C", f"{acc_vals.mean():.1f}%", "AVG ACCURACY", KPI_FILL_BLUE),
                ("E", "F", f"{acc_vals.median():.0f}%", "MEDIAN ACCURACY", KPI_FILL_TEAL),
                ("H", "I", f"{int((acc_vals >= 90).sum())}", "EXCELLENT DOCS (90%+)", KPI_FILL_GREEN),
                ("K", "L", f"{int((acc_vals < 70).sum())}", "NEEDS ATTENTION (<70%)", KPI_FILL_RED),
            ]

        for start_col, end_col, value, label, fill in kpi_data:
            ws_dash.merge_cells(f"{start_col}4:{end_col}4")
            ws_dash.merge_cells(f"{start_col}5:{end_col}5")
            val_cell = ws_dash.cell(row=4, column=ord(start_col) - 64, value=value)
            val_cell.font = KPI_NUM_FONT
            val_cell.fill = fill
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell = ws_dash.cell(row=5, column=ord(start_col) - 64, value=label)
            lbl_cell.font = KPI_LABEL_FONT
            lbl_cell.fill = fill
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            # Fill end column cells too
            ws_dash.cell(row=4, column=ord(end_col) - 64).fill = fill
            ws_dash.cell(row=5, column=ord(end_col) - 64).fill = fill

        # ================================================================
        # Write chart data tables to the Accuracy Report sheet (far right)
        # starting at column 10 (J), which is beyond the visible data
        # ================================================================
        DATA_START_COL = 10  # Column J on ws_acc

        # --- Chart Data 1: Accuracy by File Type ---
        cd1_start_row = 1
        cd1_rows = 0
        ftype_data = []
        if ftype_col_name in frame.columns and acc_col_for_dashboard in frame.columns:
            grouped = frame.groupby(ftype_col_name)[acc_col_for_dashboard]
            for ft, group in sorted(grouped, key=lambda x: x[0]):
                vals_g = pd.to_numeric(group, errors="coerce").dropna()
                if len(vals_g) > 0:
                    val_mean = round(float(vals_g.mean()), 1)
                    ft_upper = str(ft).upper()
                    ws_acc.cell(row=cd1_start_row + cd1_rows, column=DATA_START_COL, value=ft_upper)
                    ws_acc.cell(row=cd1_start_row + cd1_rows, column=DATA_START_COL + 1, value=val_mean)
                    ftype_data.append((ft_upper, val_mean))
                    cd1_rows += 1

        # --- Chart Data 2: Accuracy Distribution ---
        cd2_start_row = cd1_start_row + cd1_rows + 1
        cd2_rows = 0
        dist_data = []
        if len(acc_vals) > 0:
            dist_data = [
                ("Excellent (90-100%)", int((acc_vals >= 90).sum())),
                ("Good (80-90%)", int(((acc_vals >= 80) & (acc_vals < 90)).sum())),
                ("Fair (70-80%)", int(((acc_vals >= 70) & (acc_vals < 80)).sum())),
                ("Low (<70%)", int((acc_vals < 70).sum())),
            ]
            for label, count in dist_data:
                if count > 0:  # Only include non-zero slices
                    ws_acc.cell(row=cd2_start_row + cd2_rows, column=DATA_START_COL, value=label)
                    ws_acc.cell(row=cd2_start_row + cd2_rows, column=DATA_START_COL + 1, value=count)
                    cd2_rows += 1

        # --- Chart Data 3: Accuracy Loss Root Causes ---
        cd3_start_row = cd2_start_row + cd2_rows + 1
        cd3_rows = 0
        loss_data = []
        if loss_categories:
            # Exclude successfully read text and blank whitespace/margins
            filtered_losses = {k: v for k, v in loss_categories.items() if k not in ["Text Successfully Read", "Whitespace & Margins"]}
            sorted_loss = sorted(filtered_losses.items(), key=lambda x: -x[1])
            for cat, total_val in sorted_loss[:10]:
                val_round = round(total_val, 1)
                ws_acc.cell(row=cd3_start_row + cd3_rows, column=DATA_START_COL, value=cat)
                ws_acc.cell(row=cd3_start_row + cd3_rows, column=DATA_START_COL + 1, value=val_round)
                loss_data.append((cat, val_round))
                cd3_rows += 1

        # --- Chart Data 4: Bottom 10 Documents ---
        cd4_start_row = cd3_start_row + cd3_rows + 1
        cd4_rows = 0
        b10_list = []
        if acc_col_for_dashboard in frame.columns and "File Name" in frame.columns:
            frame_sorted = frame.copy()
            frame_sorted["_acc"] = pd.to_numeric(frame_sorted[acc_col_for_dashboard], errors="coerce")
            b10 = frame_sorted.dropna(subset=["_acc"]).nsmallest(10, "_acc")
            for _, brow in b10.iterrows():
                fname = str(brow.get("File Name", ""))
                # Truncate long file names for chart readability
                if len(fname) > 35:
                    fname = fname[:32] + "..."
                val_acc = round(float(brow.get("_acc", 0)), 1)
                ws_acc.cell(row=cd4_start_row + cd4_rows, column=DATA_START_COL, value=fname)
                ws_acc.cell(row=cd4_start_row + cd4_rows, column=DATA_START_COL + 1, value=val_acc)
                b10_list.append((fname, val_acc))
                cd4_rows += 1

        # ================================================================
        # BUILD CHARTS in memory using Matplotlib with exact styles
        # ================================================================
        plt.rcParams.update({
            'font.family': 'sans-serif',
            'font.sans-serif': ['Segoe UI', 'Arial', 'DejaVu Sans'],
            'axes.unicode_minus': False,
            'figure.facecolor': '#FFFFFF',
            'axes.facecolor': '#FFFFFF',
        })

        def fig_to_openpyxl_img(fig):
            buf = BytesIO()
            fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            img = OPImage(buf)
            plt.close(fig)
            return img

        # ---- CHART 1: Accuracy by File Type (Vertical Column Chart) ----
        if len(ftype_data) > 0:
            try:
                fig, ax = plt.subplots(figsize=(6.2, 3.8))
                labels = [x[0] for x in ftype_data]
                values = [x[1] for x in ftype_data]
                
                bars = ax.bar(labels, values, color='#1F4E79', width=0.55, edgecolor='#122E47', linewidth=0.8)
                ax.set_title("Extraction Accuracy by File Type", fontsize=12, fontweight='bold', pad=15, color='#1F4E79')
                ax.set_ylabel("Accuracy %", fontsize=10, color='#333333')
                ax.set_ylim(0, 110)
                
                for spine in ['top', 'right', 'left']:
                    ax.spines[spine].set_visible(False)
                ax.spines['bottom'].set_color('#CCCCCC')
                
                ax.grid(axis='y', linestyle='--', alpha=0.5, color='#CBD5E1')
                ax.set_axisbelow(True)
                
                for bar in bars:
                    height = bar.get_height()
                    ax.annotate(f"{height:.1f}%",
                                xy=(bar.get_x() + bar.get_width() / 2, height),
                                xytext=(0, 4),
                                textcoords="offset points",
                                ha='center', va='bottom', fontsize=8.5, fontweight='bold', color='#1F4E79')
                
                plt.tight_layout()
                ws_dash.add_image(fig_to_openpyxl_img(fig), 'B7')
            except Exception as exc:
                logger.error("Failed to generate Chart 1: %s", exc)

        # ---- CHART 2: Accuracy Quality Distribution (Donut Chart) ----
        if len(dist_data) > 0:
            try:
                dist_slices = []
                dist_labels = []
                dist_colors = []
                
                color_map = {
                    "Excellent (90-100%)": '#0D9488', # Teal Green
                    "Good (80-90%)": '#2DD4BF',      # Light Teal
                    "Fair (70-80%)": '#F59E0B',      # Soft Amber
                    "Low (<70%)": '#EF4444'          # Muted Coral
                }
                
                for label, count in dist_data:
                    if count > 0:
                        dist_slices.append(count)
                        dist_labels.append(f"{label.split(' ')[0]}\n{count} docs")
                        dist_colors.append(color_map.get(label, '#A1A1AA'))
                
                fig, ax = plt.subplots(figsize=(6.2, 3.8))
                wedges, texts, autotexts = ax.pie(
                    dist_slices, 
                    labels=dist_labels, 
                    colors=dist_colors,
                    autopct='%1.1f%%',
                    startangle=90, 
                    pctdistance=0.72,
                    wedgeprops=dict(width=0.4, edgecolor='white', linewidth=1.5),
                    textprops=dict(fontsize=8.5, color='#333333')
                )
                
                for autotext in autotexts:
                    autotext.set_fontsize(8.5)
                    autotext.set_fontweight('bold')
                    autotext.set_color('#FFFFFF')
                    
                ax.set_title("Accuracy Quality Distribution", fontsize=12, fontweight='bold', pad=15, color='#1F4E79')
                ax.axis('equal')
                plt.tight_layout()
                ws_dash.add_image(fig_to_openpyxl_img(fig), 'I7')
            except Exception as exc:
                logger.error("Failed to generate Chart 2: %s", exc)

        # ---- CHART 3: Accuracy Loss Root Causes (Horizontal Bar Chart) ----
        if len(loss_data) > 0:
            try:
                # Add a section title in Excel above the chart
                ws_dash.merge_cells("B27:O27")
                section_title = ws_dash.cell(row=27, column=2, value="⭐ KEY INSIGHT: Why Accuracy Was Lost")
                section_title.font = Font(name="Segoe UI", bold=True, size=13, color="C0504D")
                section_title.alignment = Alignment(vertical="center")

                fig, ax = plt.subplots(figsize=(13.0, 3.8))
                
                # Reverse for horizontal bar charting (bottom-up)
                categories = [x[0] for x in loss_data][::-1]
                loss_values = [x[1] for x in loss_data][::-1]
                
                bars = ax.barh(categories, loss_values, color='#C0504D', height=0.55, edgecolor='#983E3B', linewidth=0.8)
                ax.set_title("Where Accuracy Was Lost — Root Cause Breakdown", fontsize=12, fontweight='bold', pad=15, color='#C0504D')
                ax.set_xlabel("Cumulative % Impact Across All Documents", fontsize=10, color='#333333')
                
                for spine in ['top', 'right', 'left']:
                    ax.spines[spine].set_visible(False)
                ax.spines['bottom'].set_color('#CCCCCC')
                
                ax.grid(axis='x', linestyle='--', alpha=0.5, color='#CBD5E1')
                ax.set_axisbelow(True)
                
                for bar in bars:
                    width = bar.get_width()
                    ax.annotate(f"+{width:.1f}%",
                                xy=(width, bar.get_y() + bar.get_height() / 2),
                                xytext=(6, 0),
                                textcoords="offset points",
                                ha='left', va='center', fontsize=8.5, fontweight='bold', color='#C0504D')
                
                plt.tight_layout()
                ws_dash.add_image(fig_to_openpyxl_img(fig), 'B28')
            except Exception as exc:
                logger.error("Failed to generate Chart 3: %s", exc)

        # ---- CHART 4: Bottom 10 Documents (Horizontal Bar Chart) ----
        if len(b10_list) > 0:
            try:
                # Add a section title in Excel above the chart
                ws_dash.merge_cells("B47:O47")
                section_title = ws_dash.cell(row=47, column=2, value="⚠ Documents Requiring Investigation")
                section_title.font = Font(name="Segoe UI", bold=True, size=13, color="FF4444")
                section_title.alignment = Alignment(vertical="center")

                fig, ax = plt.subplots(figsize=(13.0, 4.2))
                
                # Reverse for bottom-up horizontal bar
                categories = [x[0] for x in b10_list][::-1]
                values = [x[1] for x in b10_list][::-1]
                
                bars = ax.barh(categories, values, color='#EF4444', height=0.6, edgecolor='#B91C1C', linewidth=0.8)
                ax.set_title("Bottom 10 Documents — Lowest Extraction Accuracy", fontsize=12, fontweight='bold', pad=15, color='#B91C1C')
                ax.set_xlabel("Accuracy %", fontsize=10, color='#333333')
                ax.set_xlim(0, 110)
                
                # Target threshold line at 90%
                ax.axvline(90, color='#475569', linestyle='--', linewidth=1.2, alpha=0.8)
                ax.text(90.5, 4.5, "Target: 90%", color='#475569', fontsize=9, fontweight='bold', va='center')
                
                for spine in ['top', 'right', 'left']:
                    ax.spines[spine].set_visible(False)
                ax.spines['bottom'].set_color('#CCCCCC')
                
                ax.grid(axis='x', linestyle='--', alpha=0.5, color='#CBD5E1')
                ax.set_axisbelow(True)
                
                for bar in bars:
                    width = bar.get_width()
                    ax.annotate(f"{width:.1f}%",
                                xy=(width, bar.get_y() + bar.get_height() / 2),
                                xytext=(6, 0),
                                textcoords="offset points",
                                ha='left', va='center', fontsize=8.5, fontweight='bold', color='#EF4444')
                
                plt.tight_layout()
                ws_dash.add_image(fig_to_openpyxl_img(fig), 'B48')
            except Exception as exc:
                logger.error("Failed to generate Chart 4: %s", exc)

        # Save workbook
        wb.save(str(export_file))
        return str(export_file)

    def record_taxonomy_version(
        self,
        *,
        version_id: str,
        source_file: str,
        checksum: str,
        status: str = "loaded",
        notes: str = "",
    ) -> None:
        self._ensure_schema()
        loaded_at = _utc_now_iso()

        def _insert(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                INSERT INTO taxonomy_versions (
                    version_id, source_file, checksum, loaded_at, status, notes
                ) VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(version_id) DO UPDATE SET
                    source_file=excluded.source_file,
                    checksum=excluded.checksum,
                    loaded_at=excluded.loaded_at,
                    status=excluded.status,
                    notes=excluded.notes
                """,
                (
                    str(version_id or ""),
                    str(source_file or ""),
                    str(checksum or ""),
                    loaded_at,
                    str(status or ""),
                    str(notes or ""),
                ),
            )

        self._execute_with_retry(_insert, commit=True)

    def record_tag_feedback(
        self,
        *,
        file_key: str,
        smart_id: str,
        field_name: str,
        old_value: str,
        new_value: str,
        actor: str = "user",
        reason: str = "",
        event_time: str = "",
    ) -> None:
        self._ensure_schema()
        ts = event_time or _utc_now_iso()

        def _insert(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                INSERT INTO tag_feedback (
                    file_key, smart_id, field_name, old_value, new_value, actor, event_time, reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(file_key or ""),
                    str(smart_id or ""),
                    str(field_name or ""),
                    str(old_value or ""),
                    str(new_value or ""),
                    str(actor or "user"),
                    ts,
                    str(reason or ""),
                ),
            )

        self._execute_with_retry(_insert, commit=True)

    def get_tag_feedback_weights(self) -> Dict[str, Dict[str, float]]:
        self._ensure_schema()

        def _read(conn: sqlite3.Connection) -> List[sqlite3.Row]:
            return conn.execute(
                """
                SELECT field_name, new_value, COUNT(*) AS cnt
                FROM tag_feedback
                WHERE TRIM(COALESCE(field_name, '')) != ''
                  AND TRIM(COALESCE(new_value, '')) != ''
                GROUP BY field_name, new_value
                """
            ).fetchall()

        rows = self._execute_with_retry(_read, commit=False)
        weights: Dict[str, Dict[str, float]] = {}
        for row in rows:
            field_name = str(row["field_name"] or "").strip().lower()
            new_value = str(row["new_value"] or "").strip()
            cnt = int(row["cnt"] or 0)
            if not field_name or not new_value or cnt <= 0:
                continue

            # Conservative boost to avoid overfitting; capped at +0.20.
            boost = min(0.20, 0.02 * cnt)
            weights.setdefault(field_name, {})[new_value] = round(boost, 4)
        return weights

    def create_snippet_review(
        self,
        review_id: str,
        smart_id: str,
        page_num: int,
        snippet_type: str,
        snippet_path: str,
        bounding_box: List[int],
        accuracy_impact: float,
        reviewer_role: str
    ) -> None:
        self._ensure_schema()
        def _insert(conn: sqlite3.Connection) -> None:
            conn.execute(
                """
                INSERT OR IGNORE INTO snippet_reviews (
                    review_id, smart_id, page_num, snippet_type, snippet_path,
                    bounding_box_json, accuracy_impact, reviewer_role, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending')
                """,
                (
                    review_id,
                    smart_id,
                    page_num,
                    snippet_type,
                    snippet_path,
                    json.dumps(bounding_box),
                    float(accuracy_impact),
                    reviewer_role,
                )
            )
        self._execute_with_retry(_insert, commit=True)

    def get_pending_reviews(self) -> List[Dict[str, Any]]:
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT r.*, f.file_name, f.file_path, f.extraction_accuracy, f.enhanced_accuracy
                FROM snippet_reviews r
                JOIN file_state f ON r.smart_id = f.smart_id
                WHERE r.status = 'pending'
                ORDER BY r.smart_id, r.page_num
                """
            ).fetchall()
            return [dict(row) for row in rows]
        return self._execute_with_retry(_read, commit=False)

    def update_snippet_review_status(
        self,
        review_id: str,
        status: str,
        feature_vector_path: Optional[str] = None,
        review_reason: str = "",
        reviewed_by: str = "Dashboard User"
    ) -> None:
        self._ensure_schema()
        def _update(conn: sqlite3.Connection) -> None:
            reviewed_at = _utc_now_iso()
            conn.execute(
                """
                UPDATE snippet_reviews
                SET status = ?, feature_vector_path = ?, reviewed_at = ?,
                    reviewed_by = ?, review_reason = ?
                WHERE review_id = ?
                """,
                (status, feature_vector_path, reviewed_at, reviewed_by,
                 review_reason, review_id)
            )
            # Recalculate enhanced accuracy for this document
            row = conn.execute(
                "SELECT smart_id, accuracy_impact, snippet_type FROM snippet_reviews WHERE review_id = ?",
                (review_id,)
            ).fetchone()
            if row:
                smart_id = row["smart_id"]
                snippet_type = row["snippet_type"]
                # Sum the accuracy impact of all accepted reviews for this document
                accepted_rows = conn.execute(
                    "SELECT SUM(accuracy_impact) as total_accepted FROM snippet_reviews WHERE smart_id = ? AND status = 'accepted'",
                    (smart_id,)
                ).fetchone()
                total_accepted = accepted_rows["total_accepted"] or 0.0

                # Fetch baseline accuracy
                baseline_row = conn.execute(
                    "SELECT file_key, extraction_accuracy, enhanced_accuracy, file_name FROM file_state WHERE smart_id = ?",
                    (smart_id,)
                ).fetchone()
                if baseline_row:
                    file_key = baseline_row["file_key"]
                    baseline_acc = baseline_row["extraction_accuracy"] or 0.0
                    current_enhanced = baseline_row["enhanced_accuracy"]
                    accuracy_before = current_enhanced if current_enhanced is not None else baseline_acc
                    file_name = baseline_row["file_name"] or ""

                    # Enhanced = Baseline + accepted impacts (capped at 100%)
                    enhanced_acc = min(100.0, baseline_acc + total_accepted)

                    # If all pending reviews for this document are resolved, update status to 'Approved'
                    pending_count = conn.execute(
                        "SELECT COUNT(*) as pending_cnt FROM snippet_reviews WHERE smart_id = ? AND status = 'pending'",
                        (smart_id,)
                    ).fetchone()["pending_cnt"]

                    app_status = "Approved" if total_accepted > 0 else "Full Baseline"
                    if pending_count > 0:
                        app_status = "Pending Review"

                    conn.execute(
                        """
                        UPDATE file_state
                        SET enhanced_accuracy = ?, approval_status = ?
                        WHERE file_key = ?
                        """,
                        (enhanced_acc, app_status, file_key)
                    )

                    # Record only user decision actions to keep activity log meaningful.
                    if status in ("accepted", "rejected"):
                        conn.execute(
                            """
                            INSERT INTO review_activity_log
                            (review_id, smart_id, action, actor, reason, timestamp,
                             accuracy_before, accuracy_after, snippet_type, file_name)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (review_id, smart_id, status, reviewed_by, review_reason,
                             reviewed_at, accuracy_before, enhanced_acc, snippet_type, file_name)
                        )
        self._execute_with_retry(_update, commit=True)

    def get_all_reviews_for_doc(self, smart_id: str) -> List[Dict[str, Any]]:
        """Get ALL reviews (pending, accepted, rejected) for a specific document."""
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT r.*, f.file_name, f.file_path, f.extraction_accuracy, f.enhanced_accuracy
                FROM snippet_reviews r
                JOIN file_state f ON r.smart_id = f.smart_id
                WHERE r.smart_id = ?
                ORDER BY r.page_num, r.review_id
                """,
                (smart_id,)
            ).fetchall()
            return [dict(row) for row in rows]
        return self._execute_with_retry(_read, commit=False)

    def get_review_activity_log(self, smart_id: Optional[str] = None, limit: int = 100) -> List[Dict[str, Any]]:
        """Get review activity log entries, optionally filtered by document."""
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            if smart_id:
                rows = conn.execute(
                    "SELECT * FROM review_activity_log WHERE smart_id = ? ORDER BY timestamp DESC LIMIT ?",
                    (smart_id, limit)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM review_activity_log ORDER BY timestamp DESC LIMIT ?",
                    (limit,)
                ).fetchall()
            return [dict(row) for row in rows]
        return self._execute_with_retry(_read, commit=False)

    def get_docs_with_reviews(self) -> List[Dict[str, Any]]:
        """Get all documents that have snippet reviews (any status)."""
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT f.smart_id, f.file_name, f.file_path, f.extraction_accuracy,
                       f.enhanced_accuracy, f.approval_status, f.pipeline_type,
                       COUNT(CASE WHEN r.status = 'pending' THEN 1 END) as pending_count,
                       COUNT(CASE WHEN r.status = 'accepted' THEN 1 END) as accepted_count,
                       COUNT(CASE WHEN r.status = 'rejected' THEN 1 END) as rejected_count,
                       COUNT(*) as total_snippets,
                       SUM(CASE WHEN r.status = 'pending' THEN r.accuracy_impact ELSE 0 END) as pending_impact
                FROM snippet_reviews r
                JOIN file_state f ON r.smart_id = f.smart_id
                GROUP BY r.smart_id
                ORDER BY COUNT(CASE WHEN r.status = 'pending' THEN 1 END) DESC, f.file_name
                """
            ).fetchall()
            return [dict(row) for row in rows]
        return self._execute_with_retry(_read, commit=False)

    def get_snippet_storage_stats(self) -> Dict[str, Any]:
        """Get storage statistics for snippet crop files on disk."""
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> Dict[str, Any]:
            rows = conn.execute(
                """
                SELECT r.smart_id, f.file_name, r.status,
                       COUNT(*) as snippet_count
                FROM snippet_reviews r
                LEFT JOIN file_state f ON r.smart_id = f.smart_id
                GROUP BY r.smart_id, r.status
                """
            ).fetchall()
            return [dict(row) for row in rows]
        raw = self._execute_with_retry(_read, commit=False)

        # Calculate actual disk sizes
        total_size = 0
        per_doc = {}
        for row in raw:
            sid = row.get("smart_id", "unknown")
            fname = row.get("file_name", "unknown")
            if sid not in per_doc:
                per_doc[sid] = {"file_name": fname, "total_size": 0, "file_count": 0,
                                "pending": 0, "accepted": 0, "rejected": 0}
            status = row.get("status", "unknown")
            if status in ("pending", "accepted", "rejected"):
                per_doc[sid][status] = row.get("snippet_count", 0)

        import os
        from pathlib import Path
        snippet_base = str(Path(self._config.paths.working_root).parent / "data" / "review_snippets")

        if os.path.isdir(snippet_base):
            for dirpath, _, filenames in os.walk(snippet_base):
                for fn in filenames:
                    # Storage panel tracks crop image files only.
                    if not fn.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                        continue
                    fp = os.path.join(dirpath, fn)
                    try:
                        sz = os.path.getsize(fp)
                        total_size += sz
                        # Extract smart_id from directory structure
                        rel = os.path.relpath(dirpath, snippet_base)
                        sid = rel.split(os.sep)[0] if os.sep in rel or rel != "." else rel
                        # Ensure sid exists in per_doc if found on disk
                        if sid not in per_doc:
                            per_doc[sid] = {"file_name": sid, "total_size": 0, "file_count": 0,
                                          "pending": 0, "accepted": 0, "rejected": 0}
                        per_doc[sid]["total_size"] += sz
                        per_doc[sid]["file_count"] += 1
                    except OSError:
                        pass

        return {
            "total_size_bytes": total_size,
            "total_size_mb": round(total_size / (1024 * 1024), 2),
            "per_doc": per_doc,
        }

    def purge_old_snippets(self, older_than_days: int = 30, keep_vectors: bool = True) -> Dict[str, int]:
        """Purge old accepted/rejected snippet crop files from disk.
        Keeps .npy feature vectors. Returns count of purged files and bytes freed.
        """
        import os
        from datetime import datetime, timedelta

        self._ensure_schema()
        cutoff = (datetime.utcnow() - timedelta(days=older_than_days)).isoformat() + "Z"

        def _read_purgeable(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT review_id, smart_id, snippet_path, snippet_type, status, reviewed_at
                FROM snippet_reviews
                WHERE status IN ('accepted', 'rejected')
                AND reviewed_at IS NOT NULL AND reviewed_at < ?
                """,
                (cutoff,)
            ).fetchall()
            return [dict(r) for r in rows]

        purgeable = self._execute_with_retry(_read_purgeable, commit=False)

        purged_count = 0
        bytes_freed = 0

        def _resolve_legacy_snippet_path(path_str: str) -> str:
            if not path_str:
                return ""
            if os.path.isfile(path_str):
                return path_str

            normalized = path_str.replace("\\", "/")
            marker = "/data/review_snippets/"
            if marker not in normalized:
                return path_str

            relative_part = normalized.split(marker, 1)[1]
            mapped = Path(self._config.paths.working_root).parent / "data" / "review_snippets" / Path(relative_part)
            return str(mapped)

        for item in purgeable:
            path = _resolve_legacy_snippet_path(item.get("snippet_path", ""))
            if path and os.path.isfile(path):
                try:
                    sz = os.path.getsize(path)
                    os.remove(path)
                    bytes_freed += sz
                    purged_count += 1
                except OSError:
                    pass

        # Log the purge activity
        if purged_count > 0:
            def _log_purge(conn: sqlite3.Connection) -> None:
                conn.execute(
                    """
                    INSERT INTO review_activity_log
                    (review_id, smart_id, action, actor, reason, timestamp,
                     accuracy_before, accuracy_after, snippet_type, file_name)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("system_purge", "system", "snippet_purged", "System",
                     f"Purged {purged_count} snippet files older than {older_than_days} days, freed {bytes_freed / (1024*1024):.2f} MB",
                     _utc_now_iso(), 0, 0, "all", "system_cleanup")
                )
            self._execute_with_retry(_log_purge, commit=True)

        return {"purged_count": purged_count, "bytes_freed": bytes_freed}

    def get_approved_features_for_doc(self, smart_id: str) -> List[Dict[str, Any]]:
        self._ensure_schema()
        def _read(conn: sqlite3.Connection) -> List[Dict[str, Any]]:
            rows = conn.execute(
                """
                SELECT snippet_type, feature_vector_path, bounding_box_json
                FROM snippet_reviews
                WHERE smart_id = ? AND status = 'accepted'
                """,
                (smart_id,)
            ).fetchall()
            return [dict(row) for row in rows]
        return self._execute_with_retry(_read, commit=False)


_manager_lock = threading.Lock()
_manager: Optional[_ReportingManager] = None


def _get_manager() -> _ReportingManager:
    global _manager
    if _manager is None:
        with _manager_lock:
            if _manager is None:
                _manager = _ReportingManager()
    return _manager


def record_event(event: AuditEvent) -> None:
    _get_manager().record_event(event)


def upsert_file_state(state: FileStateRow) -> None:
    _get_manager().upsert_file_state(state)


def update_accuracy_metrics(file_key: str, metrics: Dict[str, Any]) -> None:
    """Update accuracy metrics for a document by file_key."""
    _get_manager().update_accuracy_metrics(file_key, metrics)


def search_events(filter_query: str, limit: int = 50) -> List[Dict[str, Any]]:
    return _get_manager().search_events(filter_query=filter_query, limit=limit)


def get_live_feed(limit: int = 50) -> List[Dict[str, Any]]:
    return _get_manager().get_live_feed(limit=limit)


def export_state_matrix_xlsx(filters: Optional[Dict[str, Any]], out_path: str) -> str:
    return _get_manager().export_state_matrix_xlsx(filters=filters, out_path=out_path)


def record_taxonomy_version(
    *,
    version_id: str,
    source_file: str,
    checksum: str,
    status: str = "loaded",
    notes: str = "",
) -> None:
    _get_manager().record_taxonomy_version(
        version_id=version_id,
        source_file=source_file,
        checksum=checksum,
        status=status,
        notes=notes,
    )


def record_tag_feedback(
    *,
    file_key: str,
    smart_id: str,
    field_name: str,
    old_value: str,
    new_value: str,
    actor: str = "user",
    reason: str = "",
    event_time: str = "",
) -> None:
    _get_manager().record_tag_feedback(
        file_key=file_key,
        smart_id=smart_id,
        field_name=field_name,
        old_value=old_value,
        new_value=new_value,
        actor=actor,
        reason=reason,
        event_time=event_time,
    )


def get_tag_feedback_weights() -> Dict[str, Dict[str, float]]:
    return _get_manager().get_tag_feedback_weights()


def create_snippet_review(
    review_id: str,
    smart_id: str,
    page_num: int,
    snippet_type: str,
    snippet_path: str,
    bounding_box: List[int],
    accuracy_impact: float,
    reviewer_role: str
) -> None:
    _get_manager().create_snippet_review(
        review_id=review_id,
        smart_id=smart_id,
        page_num=page_num,
        snippet_type=snippet_type,
        snippet_path=snippet_path,
        bounding_box=bounding_box,
        accuracy_impact=accuracy_impact,
        reviewer_role=reviewer_role,
    )


def get_pending_reviews() -> List[Dict[str, Any]]:
    return _get_manager().get_pending_reviews()


def update_snippet_review_status(
    review_id: str,
    status: str,
    feature_vector_path: Optional[str] = None,
    review_reason: str = "",
    reviewed_by: str = "Dashboard User"
) -> None:
    _get_manager().update_snippet_review_status(
        review_id=review_id,
        status=status,
        feature_vector_path=feature_vector_path,
        review_reason=review_reason,
        reviewed_by=reviewed_by,
    )


def get_approved_features_for_doc(smart_id: str) -> List[Dict[str, Any]]:
    return _get_manager().get_approved_features_for_doc(smart_id=smart_id)


def get_all_reviews_for_doc(smart_id: str) -> List[Dict[str, Any]]:
    return _get_manager().get_all_reviews_for_doc(smart_id=smart_id)


def get_review_activity_log(smart_id: Optional[str] = None, limit: int = 100) -> List[Dict[str, Any]]:
    return _get_manager().get_review_activity_log(smart_id=smart_id, limit=limit)


def get_docs_with_reviews() -> List[Dict[str, Any]]:
    return _get_manager().get_docs_with_reviews()


def get_snippet_storage_stats() -> Dict[str, Any]:
    return _get_manager().get_snippet_storage_stats()


def purge_old_snippets(older_than_days: int = 30) -> Dict[str, int]:
    return _get_manager().purge_old_snippets(older_than_days=older_than_days)


