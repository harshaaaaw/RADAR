"""Validate the new Accuracy Dashboard sheet."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('accuracy_dashboard_v3.xlsx')

print("=" * 60)
print("SHEETS IN WORKBOOK")
print("=" * 60)
for ws in wb.worksheets:
    chart_count = len(ws._charts) if hasattr(ws, '_charts') else 0
    print(f"  {ws.title:40s}  Charts: {chart_count}")

# Check Accuracy Report has NO charts
ws_acc = [ws for ws in wb.worksheets if 'Accuracy Report' in ws.title][0]
print(f"\n{'='*60}")
print(f"ACCURACY REPORT SHEET: {len(ws_acc._charts)} charts (should be 0)")
print(f"{'='*60}")

# Check Dashboard
ws_dash = [ws for ws in wb.worksheets if 'Dashboard' in ws.title]
if ws_dash:
    ws_dash = ws_dash[0]
    print(f"\n{'='*60}")
    print(f"ACCURACY DASHBOARD: {len(ws_dash._charts)} charts")
    print(f"{'='*60}")
    
    # Print content rows
    for r in range(1, 10):
        vals = []
        for c in range(1, 16):
            v = ws_dash.cell(row=r, column=c).value
            if v is not None:
                vals.append(f"C{c}:{v}")
        if vals:
            print(f"  Row {r:2d}: {', '.join(vals)}")
    
    # KPI row
    print(f"\n--- KPI Scorecards ---")
    for r in [4, 5]:
        vals = []
        for c in range(1, 13):
            v = ws_dash.cell(row=r, column=c).value
            if v is not None:
                vals.append(f"{v}")
        print(f"  Row {r}: {' | '.join(vals)}")
    
    # Section titles
    for r in [24, 45]:
        v = ws_dash.cell(row=r, column=2).value
        if v:
            print(f"  Row {r}: {v}")
    
    # Chart anchors
    print(f"\n--- Chart Anchors ---")
    for i, chart in enumerate(ws_dash._charts):
        anchor = chart.anchor
        anchor_type = anchor.__class__.__name__
        if hasattr(anchor, '_from') and hasattr(anchor, 'to'):
            from_col = anchor._from.col
            from_row = anchor._from.row
            to_col = anchor.to.col
            to_row = anchor.to.row
            print(f"  Chart {i+1}: {anchor_type} ({from_col},{from_row}) -> ({to_col},{to_row})")
            
            # Check for overlap with other charts
            for j, other in enumerate(ws_dash._charts):
                if j <= i:
                    continue
                o_anchor = other.anchor
                if hasattr(o_anchor, '_from') and hasattr(o_anchor, 'to'):
                    o_from_row = o_anchor._from.row
                    o_to_row = o_anchor.to.row
                    if (from_row < o_to_row and to_row > o_from_row):
                        # Same column range check
                        o_from_col = o_anchor._from.col
                        o_to_col = o_anchor.to.col
                        if (from_col < o_to_col and to_col > o_from_col):
                            print(f"    ⚠️ OVERLAPS with Chart {j+1}!")
        else:
            print(f"  Chart {i+1}: {anchor_type} (OneCellAnchor - BAD!)")
    
    # Verify no overlaps
    print(f"\n--- Overlap Check ---")
    overlaps = 0
    for i in range(len(ws_dash._charts)):
        for j in range(i+1, len(ws_dash._charts)):
            a1 = ws_dash._charts[i].anchor
            a2 = ws_dash._charts[j].anchor
            if hasattr(a1, 'to') and hasattr(a2, 'to'):
                if (a1._from.row < a2.to.row and a1.to.row > a2._from.row and
                    a1._from.col < a2.to.col and a1.to.col > a2._from.col):
                    overlaps += 1
    
    if overlaps == 0:
        print("  ✅ NO OVERLAPS DETECTED")
    else:
        print(f"  ❌ {overlaps} overlaps found!")
else:
    print("\n❌ No Dashboard sheet found!")

print(f"\n{'='*60}")
print("VALIDATION COMPLETE")
print(f"{'='*60}")
