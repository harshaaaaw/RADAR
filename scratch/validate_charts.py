"""Validate chart placement and data labels in v2 export."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('accuracy_charts_v2.xlsx')

ws_acc = [ws for ws in wb.worksheets if 'Accuracy' in ws.title][0]
print(f"=== {ws_acc.title} ===")
print(f"Total rows: {ws_acc.max_row}")
print(f"Total charts: {len(ws_acc._charts)}")
print()

for i, chart in enumerate(ws_acc._charts):
    title = getattr(chart.title, 'text', str(chart.title)) if chart.title else "No title"
    # Get anchor position
    anchor = chart.anchor if hasattr(chart, 'anchor') else "unknown"
    chart_type = chart.__class__.__name__
    
    # Check data labels
    labels_info = []
    for s_idx, series in enumerate(chart.series):
        if series.dLbls:
            dl = series.dLbls
            flags = []
            if getattr(dl, 'showVal', False): flags.append("showVal")
            if getattr(dl, 'showCatName', False): flags.append("showCatName")
            if getattr(dl, 'showSerName', False): flags.append("showSerName")
            if getattr(dl, 'showPercent', False): flags.append("showPercent")
            labels_info.append(f"Series{s_idx}: {', '.join(flags)}")
    
    print(f"  Chart {i+1}: {chart_type}")
    print(f"    Title: {title}")
    print(f"    Size: {chart.width}x{chart.height}")
    print(f"    Anchor: {anchor}")
    if labels_info:
        print(f"    Labels: {'; '.join(labels_info)}")
    print()

# Check chart gallery area
print("=== CHART GALLERY ROWS ===")
for r in range(85, min(ws_acc.max_row + 1, 100)):
    vals = []
    for c in range(1, 4):
        v = ws_acc.cell(row=r, column=c).value
        if v is not None:
            vals.append(str(v)[:60])
    if vals:
        print(f"  R{r}: {' | '.join(vals)}")
