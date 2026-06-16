import openpyxl
from pathlib import Path

def main():
    path = r"C:\Users\DELL\Downloads\Metadata Mapping_ Index and Copy ticket_Box - GECIHL ILAAP V2 (1).xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    
    sheet = None
    for name in wb.sheetnames:
        if "valid" in name.lower() and "values" in name.lower():
            sheet = wb[name]
            break
            
    if not sheet:
        print("Sheet not found")
        return
        
    rows_with_desc = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        code = row[4] # Col 5 (0-indexed 4)
        desc = row[5] # Col 6 (0-indexed 5)
        if code:
            rows_with_desc.append((str(code).strip(), str(desc or '').strip()))
            
    print(f"Total rows with record type code: {len(rows_with_desc)}")
    for code, desc in rows_with_desc:
        print(f"{code:<10} : {desc}")

    wb.close()

if __name__ == '__main__':
    main()
