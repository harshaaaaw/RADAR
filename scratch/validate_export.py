"""Validate the Excel export: charts, loss formatting, sheet structure."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('accuracy_test_export.xlsx')

# === SHEET SUMMARY ===
print("=" * 60)
print("SHEET SUMMARY")
print("=" * 60)
for ws in wb.worksheets:
    charts = len(ws._charts) if hasattr(ws, '_charts') else 0
    print(f"  {ws.title}: {ws.max_row} rows, {ws.max_column} cols, {charts} charts")

# === ACCURACY REPORT DETAILS ===
ws_acc = [ws for ws in wb.worksheets if 'Accuracy' in ws.title][0]
print(f"\n{'=' * 60}")
print(f"ACCURACY REPORT SHEET: {ws_acc.title}")
print(f"{'=' * 60}")
print(f"Charts: {len(ws_acc._charts)}")
for i, chart in enumerate(ws_acc._charts):
    print(f"  Chart {i+1}: {chart.title} (type={chart.__class__.__name__})")

print(f"\nAll rows:")
for r in range(1, ws_acc.max_row + 1):
    vals = []
    for c in range(1, min(ws_acc.max_column + 1, 6)):
        v = ws_acc.cell(row=r, column=c).value
        if v is not None:
            vals.append(str(v)[:60])
    if vals:
        print(f"  R{r:3d}: {' | '.join(vals)}")

# === FORMATTED LOSS REASONS ===
print(f"\n{'=' * 60}")
print("FORMATTED ACCURACY LOSS REASONS (sample)")
print("=" * 60)
doc_sheet = [ws for ws in wb.worksheets if 'Document Data' in ws.title][0]
headers = [doc_sheet.cell(row=1, column=c).value for c in range(1, doc_sheet.max_column+1)]
loss_idx = None
ftype_idx = None
for i, h in enumerate(headers, 1):
    if h and 'Loss' in str(h):
        loss_idx = i
    if h and h == 'File Type':
        ftype_idx = i

# Show 20 diverse samples
seen_types = set()
shown = 0
for r in range(2, doc_sheet.max_row + 1):
    ftype = str(doc_sheet.cell(row=r, column=ftype_idx).value or "")
    loss = doc_sheet.cell(row=r, column=loss_idx).value
    if ftype not in seen_types or shown < 20:
        print(f"  Row {r:3d} ({ftype:5s}): {loss}")
        seen_types.add(ftype)
        shown += 1
    if shown >= 20:
        break

print(f"\n{'=' * 60}")
print("VALIDATION COMPLETE")
print("=" * 60)
