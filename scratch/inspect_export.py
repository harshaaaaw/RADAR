"""Inspect the exported Excel accuracy data."""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('accuracy_test_export.xlsx')

# --- Accuracy Report Sheet ---
ws_acc = [ws for ws in wb.worksheets if 'Accuracy' in ws.title][0]
print(f"=== {ws_acc.title} ===")
print(f"Rows: {ws_acc.max_row}, Cols: {ws_acc.max_column}")
print(f"Charts in sheet: {len(ws_acc._charts)}")
print()

for r in range(1, ws_acc.max_row + 1):
    vals = []
    for c in range(1, min(ws_acc.max_column + 1, 5)):
        v = ws_acc.cell(row=r, column=c).value
        if v is not None:
            vals.append(str(v))
    if vals:
        line = " | ".join(vals)
        print(f"  R{r}: {line}")

# --- Check accuracy loss reasons diversity ---
print("\n\n=== ACCURACY LOSS REASON ANALYSIS ===")
doc_sheet = [ws for ws in wb.worksheets if 'Document Data' in ws.title][0]
headers = [doc_sheet.cell(row=1, column=c).value for c in range(1, doc_sheet.max_column+1)]
loss_idx = None
pipe_idx = None
acc_idx = None
ftype_idx = None
for i, h in enumerate(headers, 1):
    if h and 'Loss' in str(h):
        loss_idx = i
    if h and 'Pipeline' in str(h):
        pipe_idx = i
    if h and 'Extraction Accuracy' in str(h):
        acc_idx = i
    if h and h == 'File Type':
        ftype_idx = i

# Count pipeline types
pipe_counts = {}
acc_by_pipe = {}
acc_by_ftype = {}
loss_examples = []
for r in range(2, doc_sheet.max_row + 1):
    pipe = str(doc_sheet.cell(row=r, column=pipe_idx).value or "")
    pipe_counts[pipe] = pipe_counts.get(pipe, 0) + 1
    
    acc = doc_sheet.cell(row=r, column=acc_idx).value
    if acc is not None:
        acc_by_pipe.setdefault(pipe, []).append(float(acc))
    
    ftype = str(doc_sheet.cell(row=r, column=ftype_idx).value or "")
    if acc is not None:
        acc_by_ftype.setdefault(ftype, []).append(float(acc))
    
    loss = doc_sheet.cell(row=r, column=loss_idx).value
    if loss and len(loss_examples) < 15:
        loss_examples.append((r, ftype, loss))

print("\nPipeline Distribution:")
for pipe, count in sorted(pipe_counts.items()):
    avg = sum(acc_by_pipe.get(pipe, [0])) / max(len(acc_by_pipe.get(pipe, [1])), 1)
    print(f"  {pipe}: {count} docs, avg accuracy: {avg:.2f}%")

print("\nAccuracy by File Type:")
for ft in sorted(acc_by_ftype.keys()):
    vals = acc_by_ftype[ft]
    avg = sum(vals) / max(len(vals), 1)
    print(f"  {ft}: {len(vals)} docs, avg: {avg:.2f}%, min: {min(vals):.2f}%, max: {max(vals):.2f}%")

print("\nSample Accuracy Loss Reasons:")
for r, ft, loss in loss_examples:
    print(f"  Row {r} ({ft}): {loss}")
